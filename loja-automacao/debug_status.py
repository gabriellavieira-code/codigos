#!/usr/bin/env python3
"""Debug: Mostrar todos os status das contas e dados brutos."""

import openpyxl
import sys
from pathlib import Path
from datetime import date

# Encontrar planilha DIRETAMENTE
caminho = Path.home() / "OneDrive" / "Desktop" / "ok(BMQ 2025) - GESTÃO FINANCEIRA.xlsx"
if not caminho.exists():
    print(f"Planilha não encontrada em {caminho}")
    sys.exit(1)

wb = openpyxl.load_workbook(caminho, data_only=True)
ws = wb['LP']

print("=== DADOS BRUTOS DA PLANILHA LP ===\n")
print(f"{'Linha':<5} {'Status':<15} {'Data Pag':<12} {'Descrição':<40} {'Valor':<10}")
print("-" * 85)

segunda = date(2026, 2, 23)

for i in range(10, min(ws.max_row + 1, 150)):
    desc = ws.cell(i, 3).value
    dt_pgto = ws.cell(i, 7).value
    status = ws.cell(i, 8).value
    valor = ws.cell(i, 10).value
    
    if not status and not valor:
        continue
    
    # Converter data
    dt_str = ""
    if dt_pgto:
        if hasattr(dt_pgto, 'year'):
            dt_str = dt_pgto.strftime('%d/%m/%Y')
        else:
            dt_str = str(dt_pgto)[:10]
    
    status_str = str(status or "").strip()
    desc_str = str(desc or "")[:38]
    valor_str = f"R${float(valor):,.2f}" if valor else ""
    
    # Highlight segunda-feira
    marker = ">> " if dt_pgto and hasattr(dt_pgto, 'date') and dt_pgto.date() == segunda else "   "
    
    print(f"{marker}{i:<2} {status_str:<15} {dt_str:<12} {desc_str:<40} {valor_str:<10}")
