"""
Debug: Verificar estrutura de contas a pagar e qual conta (Sicoob vs Nubank)
"""
import openpyxl
from datetime import date, timedelta

# Abrir a planilha
caminho = r"C:\Users\gaabi\OneDrive\Desktop\ok(BMQ 2025) - GESTÃO FINANCEIRA.xlsx"

try:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    print(f"✅ Planilha aberta\n")
except Exception as e:
    print(f"❌ Erro ao abrir: {e}")
    exit(1)

# Ler aba LP (Lançamentos a Pagar) com cabeçalhos
ws = wb['LP']

# Ler linha de cabeçalho
print("=== CABEÇALHOS DA ABA LP ===\n")
for col in range(1, 15):
    header = ws.cell(9, col).value
    print(f"Coluna {col}: {header}")

print("\n" + "="*60)
print("=== EXEMPLOS DE LANÇAMENTOS PARA 23/02/2026 ===\n")

# Buscar alguns exemplos
hoje = date(2026, 2, 24)
ds = hoje.weekday()
seg = hoje - timedelta(days=ds)

count = 0
for i in range(10, min(200, ws.max_row + 1)):  # Primeiras linhas
    desc = ws.cell(i, 3).value
    dt_pgto = ws.cell(i, 7).value
    status = ws.cell(i, 8).value
    valor = ws.cell(i, 10).value
    forma = ws.cell(i, 11).value  # Coluna 11 = forma de pagamento
    conta = ws.cell(i, 12).value  # Coluna 12 = conta financeira
    
    if not desc:
        continue
    
    # Verificar se é 23/02
    data_match = False
    if dt_pgto and hasattr(dt_pgto, 'year'):
        dt = dt_pgto.date() if hasattr(dt_pgto, 'date') else dt_pgto
        if dt.day == 23 and dt.month == 2 and dt.year == 2026:
            data_match = True
    
    if data_match and count < 10:  # Primeiros 10
        count += 1
        print(f"Linha {i}:")
        print(f"  Descrição: {desc}")
        print(f"  Data: {dt_pgto}")
        print(f"  Status: {status}")
        print(f"  Valor: {valor}")
        print(f"  Forma de Pagamento: {forma}")
        print(f"  Conta Financeira: {conta}")
        print()

# Agora vamos checar a soma total e dividi-la por conta
print("="*60)
print("=== TOTAL POR CONTA PARA 23/02/2026 ===\n")

total_sicoob = 0
total_nubank = 0
total_outro = 0
items_sicoob = []
items_nubank = []

for i in range(10, ws.max_row + 1):
    desc = ws.cell(i, 3).value
    dt_pgto = ws.cell(i, 7).value
    status = ws.cell(i, 8).value
    valor = ws.cell(i, 10).value
    conta = ws.cell(i, 12).value
    
    if not desc or not valor:
        continue
    
    # Normalizar data
    data_match = False
    if dt_pgto and hasattr(dt_pgto, 'year'):
        dt = dt_pgto.date() if hasattr(dt_pgto, 'date') else dt_pgto
        if dt.day == 23 and dt.month == 2 and dt.year == 2026:
            data_match = True
    
    if data_match:
        try:
            v = float(valor)
        except:
            v = 0
        
        # Classificar por conta
        conta_str = str(conta).strip() if conta else ""
        
        if "nubank" in conta_str.lower():
            total_nubank += v
            items_nubank.append((desc[:40], v))
        elif "sicoob" in conta_str.lower() or conta_str == "" or conta_str == "None":
            total_sicoob += v
            items_sicoob.append((desc[:40], v))
        else:
            total_outro += v

print(f"SICOOB (caixa principal): R${total_sicoob:,.2f}")
for desc, val in items_sicoob[:5]:
    print(f"  • {desc}: R${val:,.2f}")
if len(items_sicoob) > 5:
    print(f"  ... e +{len(items_sicoob)-5} mais")

print(f"\nNUBANK (reserva): R${total_nubank:,.2f}")
for desc, val in items_nubank[:5]:
    print(f"  • {desc}: R${val:,.2f}")
if len(items_nubank) > 5:
    print(f"  ... e +{len(items_nubank)-5} mais")

print(f"\nOUTRAS CONTAS: R${total_outro:,.2f}")

print(f"\nTOTAL GERAL: R${total_sicoob + total_nubank + total_outro:,.2f}")
