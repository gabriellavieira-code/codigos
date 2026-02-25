#!/usr/bin/env python3
"""Debug: Mostrar todos os dados de fevereiro 2026."""

import openpyxl
from datetime import date
from pathlib import Path

# Encontrar planilha DIRETAMENTE
caminho = Path.home() / "OneDrive" / "Desktop" / "ok(BMQ 2025) - GESTÃO FINANCEIRA.xlsx"
if not caminho.exists():
    print(f"Planilha não encontrada em {caminho}")
    exit(1)

wb = openpyxl.load_workbook(caminho, data_only=True)
ws = wb['LP']

print("=== CONTAS EM FEVEREIRO 2026 ===\n")
print(f"{'Linha':<5} {'Status':<15} {'Data Pag':<12} {'Descrição':<40} {'Valor':<10}")
print("-" * 85)

fev_2026_start = date(2026, 2, 1)
fev_2026_end = date(2026, 2, 28)

foundi = False
for i in range(10, min(ws.max_row + 1, 1000)):
    desc = ws.cell(i, 3).value
    dt_pgto = ws.cell(i, 7).value
    status = ws.cell(i, 8).value
    valor = ws.cell(i, 10).value
    
    if not status or not valor:
        continue
    
    # Converter data
    dt_obj = None
    dt_str = ""
    if dt_pgto:
        if hasattr(dt_pgto, 'year'):
            dt_obj = dt_pgto.date() if hasattr(dt_pgto, 'date') else dt_pgto
            dt_str = dt_obj.strftime('%d/%m/%Y')
        else:
            dt_str = str(dt_pgto)[:10]
    
    # Filtrar apenas fevereiro 2026
    if dt_obj and fev_2026_start <= dt_obj <= fev_2026_end:
        foundi = True
        status_str = str(status or "").strip()
        desc_str = str(desc or "")[:38]
        try:
            valor_str = f"R${float(valor):,.2f}" if valor else ""
        except:
            valor_str = str(valor)
        
        print(f"{i:<5} {status_str:<15} {dt_str:<12} {desc_str:<40} {valor_str:<10}")

if not foundi:
    print("❌ Nenhuma conta encontrada em fevereiro 2026!")
    print("\n=== Procurando últimas linhas da planilha ===")
    for i in range(max(10, ws.max_row - 30), ws.max_row + 1):
        desc = ws.cell(i, 3).value
        if desc:
            dt_pgto = ws.cell(i, 7).value
            status = ws.cell(i, 8).value
            valor = ws.cell(i, 10).value
            dt_str = ""
            if dt_pgto and hasattr(dt_pgto, 'year'):
                dt_str = dt_pgto.strftime('%d/%m/%Y')
            print(f"{i:<5} {str(status or ''):<15} {dt_str:<12} {str(desc)[:38]:<40}")
