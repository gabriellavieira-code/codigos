"""
Debug: Verificar dados de contas a pagar para segunda-feira 23/02
"""
import openpyxl
from datetime import date, timedelta

# Abrir a planilha
caminho_guessed = r"C:\Users\gaabi\OneDrive\Desktop\ok(BMQ 2025) - GESTÃO FINANCEIRA.xlsx"

try:
    wb = openpyxl.load_workbook(caminho_guessed, data_only=True)
    print(f"✅ Planilha aberta: {caminho_guessed}\n")
except Exception as e:
    print(f"❌ Erro ao abrir: {e}")
    exit(1)

# Ler aba LP (Lançamentos a Pagar)
ws = wb['LP']
print("=== LANÇAMENTOS COM DATA 23/02/2026 ===\n")

# Buscar segunda 23/02 na semana
hoje = date(2026, 2, 24)  # Data do relatório
ds = hoje.weekday()  # 1 = terça
seg = hoje - timedelta(days=ds)  # Volta pra 23/02 (segunda)
print(f"Semana atual: segunda {seg.strftime('%d/%m/%Y')}\n")

# Cabeçalhos
print("Lendo coluna de datas (6) e status (8)...")
print()

count = 0
encontrados = []

for i in range(10, ws.max_row + 1):
    desc = ws.cell(i, 3).value
    dt_pgto = ws.cell(i, 7).value  # Data de pagamento (PGT é coluna 7)
    status = ws.cell(i, 8).value
    valor = ws.cell(i, 10).value
    
    if not desc:
        continue
    
    # Verificar se é 23/02/2026
    data_match = False
    if dt_pgto and hasattr(dt_pgto, 'year'):
        dt = dt_pgto.date() if hasattr(dt_pgto, 'date') else dt_pgto
        if dt.day == 23 and dt.month == 2 and dt.year == 2026:
            data_match = True
    elif dt_pgto and isinstance(dt_pgto, str):
        if "23/02" in str(dt_pgto):
            data_match = True
    
    if data_match:
        count += 1
        encontrados.append({
            'linha': i,
            'desc': desc,
            'dt_pgto': dt_pgto,
            'status': status,
            'valor': valor,
        })
        print(f"Linha {i}: {desc}")
        print(f"  Data: {dt_pgto}")
        print(f"  Status: {status}")
        print(f"  Valor: {valor}")
        print()

if encontrados:
    print(f"\n✅ Encontrados {count} lançamentos para 23/02/2026\n")
    
    # Checar status
    pagas = [e for e in encontrados if e['status'] in ('Pago', 'Paga', 'Liquidado')]
    previstas = [e for e in encontrados if e['status'] == 'Previsto']
    
    print(f"  Pagas (Pago/Paga/Liquidado): {len(pagas)}")
    print(f"  Previstas: {len(previstas)}")
    print(f"  Outras: {len(encontrados) - len(pagas) - len(previstas)}")
    
    if pagas:
        print("\n💰 Contas PAGAS:")
        for p in pagas:
            print(f"   {p['desc']}: R${p['valor']:,.2f} [{p['status']}]")
    
    if previstas:
        print("\n📋 Contas PREVISTAS:")
        for pr in previstas:
            print(f"   {pr['desc']}: R${pr['valor']:,.2f}")
else:
    print(f"❌ Nenhum lançamento encontrado para 23/02/2026")

# Verificar também 24/02
print("\n" + "="*50)
print("=== LANÇAMENTOS COM DATA 24/02/2026 ===\n")
count2 = 0
for i in range(10, ws.max_row + 1):
    desc = ws.cell(i, 3).value
    dt_pgto = ws.cell(i, 7).value
    status = ws.cell(i, 8).value
    valor = ws.cell(i, 10).value
    
    if not desc:
        continue
    
    data_match = False
    if dt_pgto and hasattr(dt_pgto, 'year'):
        dt = dt_pgto.date() if hasattr(dt_pgto, 'date') else dt_pgto
        if dt.day == 24 and dt.month == 2 and dt.year == 2026:
            data_match = True
    elif dt_pgto and isinstance(dt_pgto, str):
        if "24/02" in str(dt_pgto):
            data_match = True
    
    if data_match:
        count2 += 1
        print(f"Linha {i}: {desc}")
        print(f"  Data: {dt_pgto}")
        print(f"  Status: {status}")
        print(f"  Valor: {valor}")
        print()

if count2 == 0:
    print("Nenhum lançamento encontrado para 24/02/2026")
