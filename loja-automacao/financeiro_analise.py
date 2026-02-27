"""
Módulo de Gestão Financeira — Bem Me Quer Cosméticos
Lê a planilha offline (.xlsx) e gera análises de contas a pagar, caixa, fornecedores.

Usado pelo relatorio_semanal.py
"""

import openpyxl
from datetime import date, datetime, timedelta
from collections import Counter
import os

# ── CONFIGURAÇÃO ──

# Caminho padrão da planilha no Desktop
CAMINHO_PADRAO = os.path.expanduser("~/Desktop/ok_BMQ_2025__-_GESTÃO_FINANCEIRA.xlsx")
CAMINHO_ONEDRIVE = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "ok_BMQ_2025__-_GESTÃO_FINANCEIRA.xlsx")

# Transferência semanal Nubank (CDB)
TRANSF_NUBANK_SEMANAL = 500.0

# Meta Ads — investimento mínimo semanal
META_ADS_MINIMO = 100.0

# ── REGRAS DE VIABILIDADE (estudo Bem Me Quer) ──
VENDA_MINIMA_SEGURA = 72000.0
VENDA_IDEAL = 80000.0
CMV_TETO_PCT = 35.0  # Custo mercadoria máx 35% do faturamento
META_CRESCIMENTO_PCT = 25.0  # Meta de crescimento
MARGEM_SEGURANCA_CAIXA = 5000.0
META_NUBANK = 50000.0  # Meta de reserva no Nubank

# Estrutura de custos fixos (referência mensal)
CUSTOS_FIXOS_REF = {
    "Folha + Comissões": 14000,
    "Aluguel": 5000,
    "Impostos (~9%)": None,  # Calculado dinamicamente
    "Sistemas": 370,
    "Contabilidade": 750,
    "Energia + Internet": 360,
    "Empréstimos": 4700,
    "Plano de Saúde": 231,
    "Taxas bancárias": 31,
}

# ── ALERTAS SAZONAIS (varejo cosméticos) ──
ALERTAS_SAZONAIS = {
    1: {
        "titulo": "Janeiro — Pós-festas",
        "icone": "🎆",
        "alerta": "Mês de caixa apertado! Clientes gastaram no fim do ano. "
                  "Segure compras e foque em queima de estoque. Bom mês pra liquidação.",
        "impacto_compras": "baixo",
    },
    2: {
        "titulo": "Fevereiro — Carnaval",
        "icone": "🎭",
        "alerta": "Carnaval pode aquecer maquiagem e cuidados pessoais. "
                  "Mês curto (menos dias úteis). Controle o CMV!",
        "impacto_compras": "normal",
    },
    3: {
        "titulo": "Março — Dia da Mulher",
        "icone": "💐",
        "alerta": "Dia da Mulher (08/03) aquece vendas de kits e presentes. "
                  "Prepare estoque e promoções na primeira semana! "
                  "⚠️ Ainda é pressão de caixa pós jan-fev.",
        "impacto_compras": "moderado",
    },
    4: {
        "titulo": "Abril — Preparação Dia das Mães",
        "icone": "⚠️",
        "alerta": "ATENÇÃO: Maio é Dia das Mães! Abril é mês de PREPARAR estoque. "
                  "Gaste com sabedoria — reserve caixa para marketing e pedidos de maio. "
                  "Negocie prazos longos (60/90 dias) nos pedidos de abril.",
        "impacto_compras": "alto",
    },
    5: {
        "titulo": "Maio — Dia das Mães 🎯",
        "icone": "👩",
        "alerta": "MÊS MAIS FORTE DO ANO! Dia das Mães é a data #1 do varejo de cosméticos. "
                  "Marketing agressivo, estoque preparado, equipe reforçada. "
                  "Gastos elevados com fornecedores e publicidade — mas o retorno compensa!",
        "impacto_compras": "muito alto",
    },
    6: {
        "titulo": "Junho — Dia dos Namorados + Festa Junina",
        "icone": "💕",
        "alerta": "Dia dos Namorados (12/06) aquece kits e perfumaria. "
                  "Pós Dia das Mães: avalie sobra de estoque antes de fazer pedidos grandes.",
        "impacto_compras": "moderado",
    },
    7: {
        "titulo": "Julho — Férias escolares",
        "icone": "🏖️",
        "alerta": "Mês mais fraco. Férias escolares reduzem fluxo. "
                  "Bom momento pra inventário, reorganizar estoque, negociar com fornecedores. "
                  "Segure compras!",
        "impacto_compras": "baixo",
    },
    8: {
        "titulo": "Agosto — Dia dos Pais + Preparação Primavera",
        "icone": "👨",
        "alerta": "Dia dos Pais (2º domingo) pode ter impacto leve em cosméticos masculinos. "
                  "Comece a pensar nos pedidos de setembro/outubro (Primavera/Dia das Crianças).",
        "impacto_compras": "normal",
    },
    9: {
        "titulo": "Setembro — Primavera + Preparação",
        "icone": "🌸",
        "alerta": "Início da Primavera aquece cuidados com pele e cabelo. "
                  "Prepare estoque para Outubro (Crianças) e Novembro (Black Friday).",
        "impacto_compras": "moderado",
    },
    10: {
        "titulo": "Outubro — Dia das Crianças + Halloween",
        "icone": "🎃",
        "alerta": "Dia das Crianças (12/10) pode mover kits infantis. "
                  "⚠️ Comece a negociar com fornecedores para BLACK FRIDAY (novembro). "
                  "Pedidos grandes precisam sair agora!",
        "impacto_compras": "alto",
    },
    11: {
        "titulo": "Novembro — Black Friday 🎯",
        "icone": "🖤",
        "alerta": "BLACK FRIDAY! Segunda data mais forte do ano. "
                  "Marketing pesado, descontos atrativos, estoque robusto. "
                  "Cuidado pra não vender com margem negativa. "
                  "⚠️ Dezembro vem logo — reserve caixa para pedidos de Natal!",
        "impacto_compras": "muito alto",
    },
    12: {
        "titulo": "Dezembro — Natal + Fim de Ano",
        "icone": "🎄",
        "alerta": "Natal é kits, presentes, perfumaria. Forte até dia 24. "
                  "⚠️ CUIDADO: Janeiro é mês morto! Não comprometa caixa. "
                  "Evite compras grandes após dia 15/12.",
        "impacto_compras": "alto",
    },
}


def get_alerta_sazonal():
    """Retorna alertas sazonais relevantes (mês atual + próximo)."""
    hoje = date.today()
    mes_atual = hoje.month
    mes_prox = mes_atual + 1 if mes_atual < 12 else 1

    alertas = []

    # Alerta do mês atual
    if mes_atual in ALERTAS_SAZONAIS:
        a = ALERTAS_SAZONAIS[mes_atual]
        alertas.append({**a, "mes": mes_atual, "tipo": "atual"})

    # Preview do próximo mês (se tiver impacto alto)
    if mes_prox in ALERTAS_SAZONAIS:
        prox = ALERTAS_SAZONAIS[mes_prox]
        if prox["impacto_compras"] in ("alto", "muito alto"):
            alertas.append({
                **prox, "mes": mes_prox, "tipo": "preview",
                "alerta": f"⏭️ PRÓXIMO MÊS: {prox['alerta']}"
            })

    return alertas

# Fornecedores NEGOCIÁVEIS (podem adiar boleto)
FORNECEDORES_NEGOCIAVEIS = [
    "NATURALLES",
    "HASKELL",
    "BL DISTRIBUIDORA",
    "SEGALI",
    "ALUGUEL",
    "FNE",
]

# Categorias FIXAS (não podem ser adiadas de jeito nenhum)
FIXAS_NAO_ADIA = [
    "SALÁRIO", "PRÓ LABORE", "ENERGIA", "INTERNET",
    "FGTS", "SIMPLES NACIONAL", "DARF", "INSS", "CONTABILIDADE",
    "EMPRÉSTIMO", "SICOOB - EMPRÉSTIMO", "PLANO DE SAÚDE",
    "TAXA DE BANCO", "SISTEMA",
]


def encontrar_planilha():
    """Encontra a planilha financeira no Desktop."""
    for caminho in [CAMINHO_PADRAO, CAMINHO_ONEDRIVE]:
        if os.path.exists(caminho):
            return caminho
    for pasta in [os.path.expanduser("~/Desktop"),
                  os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")]:
        if os.path.exists(pasta):
            for f in os.listdir(pasta):
                if f.endswith(".xlsx") and ("GEST" in f.upper() or "FINANC" in f.upper()):
                    return os.path.join(pasta, f)
    return None


def eh_fixa(descricao):
    """Verifica se a conta é fixa (não pode ser adiada)."""
    desc_upper = (descricao or "").upper()
    for termo in FIXAS_NAO_ADIA:
        if termo in desc_upper:
            return True
    return False


def eh_negociavel(fornecedor, descricao):
    """Verifica se o fornecedor/conta é negociável."""
    texto = f"{fornecedor} {descricao}".upper()
    for neg in FORNECEDORES_NEGOCIAVEIS:
        if neg in texto:
            return True
    return False


def empurrar_para_dia_util(dt):
    """Se cai em sábado ou domingo, empurra pra segunda seguinte."""
    if dt is None:
        return None
    ds = dt.weekday()
    if ds == 5:  # Sábado → segunda
        return dt + timedelta(days=2)
    elif ds == 6:  # Domingo → segunda
        return dt + timedelta(days=1)
    return dt


def ler_contas_a_pagar(wb):
    """Lê todos os lançamentos da aba LP."""
    ws = wb['LP']
    contas = []
    for i in range(10, ws.max_row + 1):
        desc = ws.cell(i, 3).value
        cat = ws.cell(i, 4).value
        classif = ws.cell(i, 5).value
        dt_comp = ws.cell(i, 6).value
        dt_pgto = ws.cell(i, 7).value
        status = ws.cell(i, 8).value
        forn = ws.cell(i, 9).value
        valor = ws.cell(i, 10).value
        forma = ws.cell(i, 11).value
        conta_fin = ws.cell(i, 12).value
        centro = ws.cell(i, 13).value

        if not desc and not valor:
            continue
        if not status:
            continue

        try:
            v = float(valor) if valor else 0
        except (ValueError, TypeError):
            v = 0

        # Normalizar data
        dt = None
        if dt_pgto and hasattr(dt_pgto, 'year'):
            dt = dt_pgto.date() if hasattr(dt_pgto, 'date') else dt_pgto
        elif dt_pgto and isinstance(dt_pgto, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    dt = datetime.strptime(dt_pgto.strip(), fmt).date()
                    break
                except ValueError:
                    continue

        nome_forn = str(forn).strip() if forn else ""
        if not nome_forn:
            nome_forn = str(desc).strip()[:50] if desc else "SEM IDENTIFICAÇÃO"

        # Data de pagamento real (empurrando fim de semana)
        dt_original = dt
        dt_util = empurrar_para_dia_util(dt)

        contas.append({
            "descricao": str(desc or "").strip(),
            "categoria": str(cat or "").strip(),
            "classificacao": str(classif or "").strip(),
            "data_original": dt_original,
            "data_pagamento": dt_util,  # Data real de pagamento (dia útil)
            "status": str(status).strip(),
            "fornecedor": nome_forn,
            "valor": v,
            "forma_pagamento": str(forma or "").strip(),
            "conta_financeira": str(conta_fin or "").strip(),
            "centro_custo": str(centro or "").strip(),
            "fixa": eh_fixa(str(desc or "")),
            "negociavel": eh_negociavel(nome_forn, str(desc or "")),
        })

    return contas


def ler_saldo_caixa(wb):
    """Lê o saldo acumulado mais recente da aba FCD."""
    ws = wb['FCD']
    ultimo_saldo = 0
    ultima_data = None

    for i in range(14, ws.max_row + 1):
        dia = ws.cell(i, 4).value
        acum = ws.cell(i, 8).value

        if dia and hasattr(dia, 'year'):
            dt = dia.date() if hasattr(dia, 'date') else dia
            if acum is not None:
                try:
                    ultimo_saldo = float(acum)
                    ultima_data = dt
                except (ValueError, TypeError):
                    pass

    return {"saldo": round(ultimo_saldo, 2), "data": ultima_data}


def ler_saldos_contas(wb):
    """Lê saldos de todas as contas financeiras (AVAH ou OUTROS)."""
    # Tentar aba AVAH primeiro (mais atualizada)
    saldos = {"sicoob": 0, "nubank": 0, "caixa_loja": 0}

    # Tentar aba AVAH primeiro (mais atualizada), depois OUTROS
    for aba_nome in ['AVAH', 'OUTROS']:
        if aba_nome not in wb.sheetnames:
            continue
        ws = wb[aba_nome]
        for i in range(1, min(50, ws.max_row + 1)):
            for j in range(1, min(15, ws.max_column + 1)):
                celula = str(ws.cell(i, j).value or "").strip().upper()
                if "NUBANK" in celula:
                    val = ws.cell(i, j + 1).value
                    try:
                        v = float(val)
                        if v is not None and saldos["nubank"] == 0:  # Só pega o primeiro encontrado
                            saldos["nubank"] = round(v, 2)
                    except (ValueError, TypeError):
                        pass
                elif "SICOOB" in celula and "NUBANK" not in celula:
                    val = ws.cell(i, j + 1).value
                    try:
                        v = float(val)
                        if v is not None and saldos["sicoob"] == 0:
                            saldos["sicoob"] = round(v, 2)
                    except (ValueError, TypeError):
                        pass
                elif "CAIXA LOJA" in celula:
                    val = ws.cell(i, j + 1).value
                    try:
                        v = float(val)
                        if v is not None and saldos["caixa_loja"] == 0:
                            saldos["caixa_loja"] = round(v, 2)
                    except (ValueError, TypeError):
                        pass

    saldos["total_geral"] = round(saldos["sicoob"] + saldos["nubank"] + saldos["caixa_loja"], 2)
    saldos["nubank_falta_meta"] = round(max(0, META_NUBANK - saldos["nubank"]), 2)
    saldos["nubank_pct_meta"] = round(saldos["nubank"] / META_NUBANK * 100, 1) if META_NUBANK > 0 else 0

    return saldos


def avaliar_uso_nubank(saldos, saldo_apos_contas_com_receita):
    """Avalia se deve usar reserva do Nubank — só em último caso."""
    nubank = saldos["nubank"]

    if saldo_apos_contas_com_receita >= 1000:
        return {
            "usar": False,
            "tipo": "ok",
            "mensagem": f"💜 Nubank preservado! Saldo: R${nubank:,.2f} "
                        f"({saldos['nubank_pct_meta']:.0f}% da meta de R${META_NUBANK:,.2f}). "
                        f"Faltam R${saldos['nubank_falta_meta']:,.2f} pra meta.",
        }
    elif saldo_apos_contas_com_receita >= -2000:
        return {
            "usar": False,
            "tipo": "atencao",
            "mensagem": f"⚠️ Semana apertada, mas NÃO mexa no Nubank. "
                        f"Negocie adiamentos primeiro. "
                        f"Nubank: R${nubank:,.2f} ({saldos['nubank_pct_meta']:.0f}% da meta).",
        }
    else:
        deficit = abs(saldo_apos_contas_com_receita)
        return {
            "usar": True,
            "tipo": "critico",
            "valor_sugerido": round(min(deficit, nubank * 0.3), 2),  # Máx 30% do Nubank
            "mensagem": f"🚨 Caixa insuficiente mesmo com adiamentos. "
                        f"Se necessário, use ATÉ R${min(deficit, nubank * 0.3):,.2f} do Nubank "
                        f"(máx 30% da reserva). Nubank atual: R${nubank:,.2f}. "
                        f"⚠️ Reponha na primeira oportunidade!",
        }


def gerar_contas_semana(contas, data_referencia=None):
    """Gera relatório de contas a pagar da semana (seg a sex), reconhecendo contas já pagas."""
    if data_referencia is None:
        data_referencia = date.today()

    # Calcular segunda a sexta da semana corrente ou próxima
    ds = data_referencia.weekday()
    if ds == 0:  # Segunda
        seg = data_referencia
    elif ds <= 4:  # Ter a Sex — semana atual
        seg = data_referencia - timedelta(days=ds)
    else:  # Sáb/Dom — próxima semana
        seg = data_referencia + timedelta(days=(7 - ds))

    sex = seg + timedelta(days=4)

    # Filtrar contas da semana (seg a sex) — PREVISTAS ou PAGAS
    # Nota: contas de sáb/dom já foram empurradas pra segunda
    previstas = [c for c in contas if c["status"] in ("Previsto", "Pago", "Paga", "Liquidado", "Realizado") 
                 and c["data_pagamento"] and seg <= c["data_pagamento"] <= sex]
    
    # Apenas contas não-pagas entram no total de "contas a pagar"
    nao_pagas = [c for c in previstas if c["status"] in ("Previsto",)]
    pagas = [c for c in previstas if c["status"] in ("Pago", "Paga", "Liquidado", "Realizado")]

    # Agrupar por dia
    nomes_dias = {0: "SEGUNDA-FEIRA", 1: "TERÇA-FEIRA", 2: "QUARTA-FEIRA",
                  3: "QUINTA-FEIRA", 4: "SEXTA-FEIRA"}

    dias = {}
    for c in previstas:
        chave = c["data_pagamento"]
        if chave not in dias:
            dias[chave] = {"nome_dia": nomes_dias.get(chave.weekday(), "?"),
                           "data": chave, "contas": [], "pagas": [], "total": 0, "total_realizado": 0,
                           "total_sicoob": 0, "total_nubank": 0,
                           "passou": chave < date.today()}
        
        # Separar pagas de pendentes E identificar tipo de conta (Sicoob vs Nubank)
        conta_financeira = str(c.get("conta_financeira", "")).strip().lower()
        is_nubank = "nubank" in conta_financeira
        
        if c["status"] in ("Pago", "Paga", "Liquidado", "Realizado"):
            dias[chave]["pagas"].append(c)
            dias[chave]["total_realizado"] += c["valor"]
            if is_nubank:
                dias[chave]["total_nubank"] += c["valor"]
            else:
                dias[chave]["total_sicoob"] += c["valor"]
        else:
            dias[chave]["contas"].append(c)
            dias[chave]["total"] += c["valor"]
            dias[chave]["total_realizado"] += c["valor"]
            if is_nubank:
                dias[chave]["total_nubank"] += c["valor"]
            else:
                dias[chave]["total_sicoob"] += c["valor"]

    # Preencher dias vazios (terça zerada, etc)
    for i in range(5):
        d = seg + timedelta(days=i)
        if d not in dias:
            dias[d] = {"nome_dia": nomes_dias.get(d.weekday(), "?"),
                       "data": d, "contas": [], "pagas": [], "total": 0, "total_realizado": 0,
                       "total_sicoob": 0, "total_nubank": 0,
                       "passou": d <= date.today()}

    dias_ordenados = sorted(dias.values(), key=lambda x: x["data"])

    # Totais — APENAS CONTAS NÃO PAGAS
    total_semana = sum(d["total"] for d in dias_ordenados)
    total_pagas = sum(sum(c["valor"] for c in d["pagas"]) for d in dias_ordenados)
    fixas = sum(c["valor"] for c in nao_pagas if c["fixa"])
    negociaveis = sum(c["valor"] for c in nao_pagas if c["negociavel"])
    outras = total_semana - fixas - negociaveis

    # Dia mais pesado
    dia_mais_pesado = max(dias_ordenados, key=lambda x: x["total"]) if dias_ordenados else None

    # Fornecedores múltiplos na semana
    forn_semana = Counter()
    forn_valor = {}
    for c in nao_pagas:  # Apenas não-pagas
        forn_semana[c["fornecedor"]] += 1
        forn_valor[c["fornecedor"]] = forn_valor.get(c["fornecedor"], 0) + c["valor"]
    fornecedores_multiplos = {f: {"count": c, "valor": forn_valor[f]}
                              for f, c in forn_semana.items() if c >= 2}

    # ── Texto copiável (formato que Gabi já envia) ──
    texto_partes = []
    texto_partes.append("CONTAS A PAGAR NA SEMANA")
    texto_partes.append(f"Semana {seg.strftime('%d/%m')} a {sex.strftime('%d/%m/%Y')}")
    texto_partes.append("_________")

    for d in dias_ordenados:
        texto_partes.append(f"{d['nome_dia']} {d['data'].strftime('%d/%m')}")
        
        # Contas pendentes
        if not d["contas"]:
            texto_partes.append("Zerado")
        else:
            for c in sorted(d["contas"], key=lambda x: x["valor"], reverse=True):
                conta_tag = f" [{c['conta_financeira']}]" if c["conta_financeira"] not in ("SICOOB", "", "None") else ""
                neg_tag = " ✅NEGOCIÁVEL" if c["negociavel"] else ""
                forn = c.get("fornecedor") or ""
                forma = c.get("forma_pagamento") or ""
                texto_partes.append(
                    f"  {c['data_original'].strftime('%d/%m/%Y') if c['data_original'] else '?'}\t"
                    f"{c['descricao'][:55]}\t"
                    f"Previsto\t"
                    f"{forn}\t"
                    f"R$ {c['valor']:,.2f}{conta_tag}{neg_tag}\t"
                    f"{forma}"
                )
        
        # Contas pagas (destacar)
        if d["pagas"]:
            texto_partes.append("✅ JÁ PAGO:")
            for c in sorted(d["pagas"], key=lambda x: x["valor"], reverse=True):
                forn = c.get("fornecedor") or ""
                forma = c.get("forma_pagamento") or ""
                texto_partes.append(
                    f"  {c['descricao'][:55]}\t"
                    f"{forn}\t"
                    f"R$ {c['valor']:,.2f}\t"
                    f"{forma}"
                )
        
        texto_partes.append(f"TOTAL: R$ {d['total']:,.2f}")
        if d["pagas"]:
            total_pagas_dia = sum(c["valor"] for c in d["pagas"])
            texto_partes.append(f"Já pago: R$ {total_pagas_dia:,.2f}")
        texto_partes.append("_________")

    texto_partes.append(f"\nA PAGAR: R$ {total_semana:,.2f}")
    texto_partes.append(f"JÁ PAGO: R$ {total_pagas:,.2f}")
    if negociaveis > 0:
        texto_partes.append(f"  → Negociáveis: R$ {negociaveis:,.2f}")

    texto_copiavel = "\n".join(texto_partes)

    return {
        "inicio": seg,
        "fim": sex,
        "dias": dias_ordenados,
        "total_semana": round(total_semana, 2),  # Apenas não-pagas
        "total_pagas": round(total_pagas, 2),  # Contas já pagas
        "fixas": round(fixas, 2),
        "negociaveis": round(negociaveis, 2),
        "outras": round(outras, 2),
        "dia_mais_pesado": dia_mais_pesado,
        "fornecedores_multiplos": fornecedores_multiplos,
        "texto_copiavel": texto_copiavel,
        "qtd_contas": len(nao_pagas),
    }


def gerar_insights_semana(sem, saldo_info, contas):
    """Gera insights e recomendações inteligentes para a semana."""
    insights = []
    saldo = saldo_info["saldo"]
    total = sem["total_semana"]

    # 1) Caixa cobre ou não?
    sobra = saldo - total
    if sobra < 0:
        insights.append({
            "tipo": "critico",
            "icone": "🚨",
            "texto": f"O caixa NÃO cobre as contas da semana. Faltam R${abs(sobra):,.2f}. "
                     f"Priorize as fixas (R${sem['fixas']:,.2f}) e negocie adiamento das negociáveis (R${sem['negociaveis']:,.2f})."
        })
    elif sobra < 3000:
        insights.append({
            "tipo": "atencao",
            "icone": "⚠️",
            "texto": f"Semana apertada! Após pagar tudo, sobram apenas R${sobra:,.2f} no caixa. Evite compras extras."
        })
    else:
        insights.append({
            "tipo": "ok",
            "icone": "✅",
            "texto": f"Caixa confortável! Após as contas, sobram R${sobra:,.2f}."
        })

    # 2) Dia mais pesado
    dmp = sem.get("dia_mais_pesado")
    if dmp and dmp["total"] > 5000:
        insights.append({
            "tipo": "atencao",
            "icone": "📅",
            "texto": f"{dmp['nome_dia']} ({dmp['data'].strftime('%d/%m')}) é o dia mais pesado: R${dmp['total']:,.2f}. "
                     f"Se precisar aliviar, veja se alguma conta desse dia é negociável."
        })

    # 3) Concentração de fornecedor
    for f, info in sem.get("fornecedores_multiplos", {}).items():
        insights.append({
            "tipo": "info",
            "icone": "📌",
            "texto": f"{f} aparece {info['count']}x essa semana, totalizando R${info['valor']:,.2f}. "
                     f"Nas próximas compras, tente negociar prazos escalonados para não concentrar."
        })

    # 4) Análise de prazos — muitos boletos curtos?
    prazos_curtos = [c for c in contas if c["status"] == "Previsto"
                     and c["data_pagamento"] and c.get("data_original")
                     and sem["inicio"] <= c["data_pagamento"] <= sem["fim"]
                     and not c["fixa"]]
    if len(prazos_curtos) > 8:
        insights.append({
            "tipo": "estrategia",
            "icone": "💡",
            "texto": f"Há {len(prazos_curtos)} boletos de compras essa semana. Para as próximas negociações, "
                     f"priorize prazos de 30/60/90 dias e tente espaçar os vencimentos ao longo do mês."
        })

    # 5) Sugestão geral de melhoria
    if sem["negociaveis"] > sem["fixas"]:
        insights.append({
            "tipo": "estrategia",
            "icone": "🤝",
            "texto": f"R${sem['negociaveis']:,.2f} são de fornecedores negociáveis — "
                     f"isso é {sem['negociaveis']/total*100:.0f}% do total. "
                     f"Na reunião de compras, avalie redistribuir esses vencimentos pro meio/fim do mês."
        })

    return insights


def calcular_transferencia_nubank(saldo_apos_contas):
    """Decide se deve transferir R$500 pro Nubank essa semana."""
    if saldo_apos_contas >= 5000:
        return {
            "transferir": True,
            "valor": TRANSF_NUBANK_SEMANAL,
            "dia": "segunda-feira",
            "mensagem": f"✅ Pode transferir R${TRANSF_NUBANK_SEMANAL:,.2f} pro Nubank normalmente. "
                        f"Após contas + transferência, sobram R${saldo_apos_contas - TRANSF_NUBANK_SEMANAL:,.2f}.",
            "tipo": "ok",
        }
    elif saldo_apos_contas >= 2000:
        return {
            "transferir": True,
            "valor": TRANSF_NUBANK_SEMANAL,
            "dia": "quarta ou quinta",
            "mensagem": f"⚠️ Semana apertada. Transfira os R${TRANSF_NUBANK_SEMANAL:,.2f} só na quarta ou quinta, "
                        f"depois de confirmar que as contas de segunda e terça foram pagas.",
            "tipo": "atencao",
        }
    else:
        return {
            "transferir": False,
            "valor": 0,
            "dia": None,
            "mensagem": f"🚨 Semana muito apertada — NÃO transfira essa semana. "
                        f"Lembre que esse valor vai se acumulando: "
                        f"na próxima semana transfira R${TRANSF_NUBANK_SEMANAL * 2:,.2f} (2 semanas).",
            "tipo": "critico",
        }


def calcular_meta_ads(saldo_apos_contas_e_transf):
    """Decide se investe em tráfego pago essa semana."""
    if saldo_apos_contas_e_transf >= 3000:
        return {
            "investir": True,
            "valor": META_ADS_MINIMO,
            "mensagem": f"🚀 Que tal colocar R${META_ADS_MINIMO:,.2f} em tráfego essa semana? "
                        f"Manter a presença constante traz resultados acumulados!",
            "tipo": "ok",
        }
    elif saldo_apos_contas_e_transf >= 1500:
        return {
            "investir": True,
            "valor": META_ADS_MINIMO,
            "mensagem": f"📊 Dá pra manter os R${META_ADS_MINIMO:,.2f} de tráfego, mas fique de olho no caixa.",
            "tipo": "atencao",
        }
    else:
        return {
            "investir": False,
            "valor": 0,
            "mensagem": "⏸️ Essa semana está apertada — vamos segurar as campanhas. "
                        "Retome assim que o caixa melhorar!",
            "tipo": "critico",
        }


def gerar_projecao_caixa(saldo_info, contas):
    """Projeta saldo para o fim do mês baseado nas contas previstas."""
    hoje = date.today()
    if hoje.month < 12:
        fim_mes = date(hoje.year, hoje.month + 1, 1) - timedelta(days=1)
    else:
        fim_mes = date(hoje.year, 12, 31)

    previstas_mes = [c for c in contas if c["status"] == "Previsto"
                     and c["data_pagamento"]
                     and hoje <= c["data_pagamento"] <= fim_mes]

    total_a_pagar = sum(c["valor"] for c in previstas_mes)
    saldo_projetado = saldo_info["saldo"] - total_a_pagar

    por_categoria = {}
    for c in previstas_mes:
        cat = c["categoria"] if c["categoria"] else "Outros"
        por_categoria[cat] = por_categoria.get(cat, 0) + c["valor"]

    return {
        "saldo_atual": saldo_info["saldo"],
        "saldo_data": saldo_info["data"],
        "total_a_pagar_mes": round(total_a_pagar, 2),
        "saldo_projetado": round(saldo_projetado, 2),
        "qtd_contas_restantes": len(previstas_mes),
        "por_categoria": {k: round(v, 2) for k, v in
                          sorted(por_categoria.items(), key=lambda x: x[1], reverse=True)},
    }


def gerar_analise_fornecedores(contas, meses=12):
    """Gera análise de fornecedores dos últimos N meses (para fechamento)."""
    corte = date.today() - timedelta(days=meses * 30)
    realizados = [c for c in contas if c["status"] == "Realizado"
                  and c["data_pagamento"] and c["data_pagamento"] >= corte]

    fornecedores = {}
    for c in realizados:
        f = c["fornecedor"]
        if f not in fornecedores:
            fornecedores[f] = {
                "total": 0, "count": 0, "valores": [],
                "formas": Counter(), "categoria": c["categoria"],
            }
        fornecedores[f]["total"] += c["valor"]
        fornecedores[f]["count"] += 1
        fornecedores[f]["valores"].append(c["valor"])
        fornecedores[f]["formas"][c["forma_pagamento"]] += 1

    ranking = []
    for f, d in fornecedores.items():
        media = d["total"] / d["count"] if d["count"] > 0 else 0
        forma_principal = d["formas"].most_common(1)[0][0] if d["formas"] else "?"
        mensal = d["total"] / meses
        ranking.append({
            "fornecedor": f,
            "total": round(d["total"], 2),
            "count": d["count"],
            "media": round(media, 2),
            "mensal": round(mensal, 2),
            "forma_principal": forma_principal,
            "categoria": d["categoria"],
        })

    ranking.sort(key=lambda x: x["total"], reverse=True)
    return ranking


def gerar_limite_compras(saldo_info, contas):
    """Calcula quanto pode ser liberado para compras no próximo mês."""
    hoje = date.today()
    if hoje.month < 12:
        prox_mes = hoje.month + 1
        prox_ano = hoje.year
    else:
        prox_mes = 1
        prox_ano = hoje.year + 1

    try:
        inicio_prox = date(prox_ano, prox_mes, 1)
        if prox_mes < 12:
            fim_prox = date(prox_ano, prox_mes + 1, 1) - timedelta(days=1)
        else:
            fim_prox = date(prox_ano, 12, 31)
    except ValueError:
        fim_prox = date(prox_ano, prox_mes, 28)

    compromissos_prox = [c for c in contas if c["status"] == "Previsto"
                         and c["data_pagamento"]
                         and inicio_prox <= c["data_pagamento"] <= fim_prox]
    total_compromissos = sum(c["valor"] for c in compromissos_prox)
    fixos_prox = sum(c["valor"] for c in compromissos_prox if c["fixa"])
    compras_ja = sum(c["valor"] for c in compromissos_prox if not c["fixa"])

    margem_seguranca = MARGEM_SEGURANCA_CAIXA
    limite_sugerido = max(0, saldo_info["saldo"] - total_compromissos - margem_seguranca)

    return {
        "mes_referencia": f"{prox_mes:02d}/{prox_ano}",
        "compromissos_fixos": round(fixos_prox, 2),
        "compras_ja_previstas": round(compras_ja, 2),
        "total_compromissos": round(total_compromissos, 2),
        "margem_seguranca": margem_seguranca,
        "limite_sugerido": round(limite_sugerido, 2),
        "saldo_base": saldo_info["saldo"],
    }


def calcular_teto_diario_pagamento(receita_info, semana_obj):
    """
    Calcula o teto recomendado de pagamento por dia.
    Baseado na receita média diária × fator de segurança (50-60%).
    
    Args:
        receita_info: Dict com "media_diaria_ap" (receita média/dia)
        semana_obj: Dict com dados da semana (dias, total_semana)
    
    Returns:
        Dict com teto diário, análise por dia e recomendações
    """
    media_diaria = receita_info.get("media_diaria_ap", 0)
    if media_diaria <= 0:
        return {"teto_diario": 0, "dias": [], "recomendacao": "Sem histórico de receita"}
    
    # Teto de 55% da média diária (margem de segurança: 45% fica no caixa)
    teto_diario = media_diaria * 0.55
    
    # Analisar cada dia da semana
    dias_analise = []
    for d in semana_obj.get("dias", []):
        dia_nome = d["nome_dia"]
        dia_data = d["data"]
        # ✅ USAR total_sicoob (APENAS caixa operacional de Sicoob, não Nubank que é reserva)
        # Se total_sicoob existe, usar; senão usar total_realizado por backwards compatibility
        total_dia = d.get("total_sicoob", 0) if d.get("total_sicoob", 0) > 0 or d.get("total_nubank", 0) == 0 else d.get("total_realizado", d["total"])
        
        # Adicionar badge se já passou e foi realizado
        passou = d.get("passou", False)
        status_badge = " ✅ Já realizado" if passou and d.get("total_realizado", 0) > 0 else ""
        
        percentual = (total_dia / teto_diario * 100) if teto_diario > 0 else 0
        
        # Classificar viabilidade
        if total_dia == 0:
            viabilidade = "ok"
            icone = "✅"
            texto = "Zerado"
        elif percentual <= 80:
            viabilidade = "ok"
            icone = "✅"
            texto = f"Confortável ({percentual:.0f}% do teto)"
        elif percentual <= 100:
            viabilidade = "atencao"
            icone = "⚠️"
            texto = f"No limite ({percentual:.0f}% do teto)"
        else:
            viabilidade = "critico"
            icone = "🚨"
            texto = f"Acima do teto ({percentual:.0f}% do teto) — negocie adiamentos!"
        
        dias_analise.append({
            "dia": dia_nome,
            "data": dia_data,
            "valor": total_dia,
            "valor_nubank": d.get("total_nubank", 0),  # Mostrar Nubank separado
            "percentual": percentual,
            "viabilidade": viabilidade,
            "icone": icone,
            "texto": texto + status_badge,  # ✅ Adicionar aviso de realizado
            "passou": passou,
        })
    
    # Recomendação geral
    total_semana = semana_obj.get("total_semana", 0)
    pct_semana = (total_semana / (teto_diario * 5) * 100) if teto_diario > 0 else 0
    
    if pct_semana > 110:
        recomendacao = f"⚠️ Semana ACIMA da capacidade de pagamento ({pct_semana:.0f}%). Negocie adiamentos ou prazos maiores."
    elif pct_semana > 100:
        recomendacao = f"⚠️ Semana no limite ({pct_semana:.0f}%). Avalie redistribuir alguns pagamentos."
    else:
        recomendacao = f"✅ Semana OK ({pct_semana:.0f}% de ocupação). Há espaço para novos pedidos."
    
    return {
        "teto_diario_recomendado": round(teto_diario, 2),
        "media_diaria_receita": round(media_diaria, 2),
        "fator_seguranca": 0.55,  # 55% — deixa 45% de margem
        "dias": dias_analise,
        "total_semana": round(total_semana, 2),
        "ocupacao_semana_pct": round(pct_semana, 1),
        "recomendacao": recomendacao,
    }


def prever_receita_semanal(wb):
    """Estima receita baseada no FCD do mesmo mês do ano anterior (referência comparativa).
    Filtra entradas atípicas (empréstimos, aportes) pra não distorcer."""
    ws = wb['FCD']
    hoje = date.today()
    mes_atual = hoje.month
    ano_passado = hoje.year - 1

    # Entradas do mesmo mês no ano anterior
    entradas_ap = []
    for i in range(14, ws.max_row + 1):
        dia = ws.cell(i, 4).value
        ent = ws.cell(i, 5).value
        if not dia or not hasattr(dia, 'year'):
            continue
        dt = dia.date() if hasattr(dia, 'date') else dia
        if dt.year == ano_passado and dt.month == mes_atual:
            try:
                v = float(ent) if ent else 0
            except (ValueError, TypeError):
                v = 0
            if v > 0:
                entradas_ap.append(v)

    # Filtrar outliers (empréstimos/aportes): remover valores > 3x a mediana
    if entradas_ap:
        entradas_ap_sorted = sorted(entradas_ap)
        mediana = entradas_ap_sorted[len(entradas_ap_sorted) // 2]
        limite_outlier = mediana * 3
        entradas_filtradas = [v for v in entradas_ap if v <= limite_outlier]
    else:
        entradas_filtradas = []

    total_ap = sum(entradas_filtradas)
    dias_ap = len(entradas_filtradas)
    media_diaria_ap = total_ap / max(dias_ap, 1)

    # Entradas do mês atual até agora (mesma filtragem)
    entradas_atual = []
    for i in range(14, ws.max_row + 1):
        dia = ws.cell(i, 4).value
        ent = ws.cell(i, 5).value
        if not dia or not hasattr(dia, 'year'):
            continue
        dt = dia.date() if hasattr(dia, 'date') else dia
        if dt.year == hoje.year and dt.month == hoje.month and dt <= hoje:
            try:
                v = float(ent) if ent else 0
            except (ValueError, TypeError):
                v = 0
            if v > 0:
                entradas_atual.append(v)

    # Filtrar outliers do mês atual também (Pronampe, etc)
    if entradas_atual:
        atual_sorted = sorted(entradas_atual)
        mediana_atual = atual_sorted[len(atual_sorted) // 2]
        limite_atual = mediana_atual * 3
        entradas_atual_filtradas = [v for v in entradas_atual if v <= limite_atual]
        outliers_removidos = sum(v for v in entradas_atual if v > limite_atual)
    else:
        entradas_atual_filtradas = []
        outliers_removidos = 0

    entradas_mes = sum(entradas_atual_filtradas)
    dias_com_entrada = len(entradas_atual_filtradas)

    # Dias restantes no mês
    if hoje.month < 12:
        fim_mes = date(hoje.year, hoje.month + 1, 1) - timedelta(days=1)
    else:
        fim_mes = date(hoje.year, 12, 31)

    dias_rest = 0
    for d in range(1, (fim_mes - hoje).days + 1):
        dt = hoje + timedelta(days=d)
        if dt.weekday() < 6:
            dias_rest += 1

    # Previsão conservadora: usa média do ano passado
    previsao_entrada_restante = media_diaria_ap * dias_rest

    nomes_meses = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
                   7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}

    return {
        "ref_ano_passado": round(total_ap, 2),
        "ref_mes": f"{nomes_meses.get(mes_atual, '?')}/{ano_passado}",
        "media_diaria_ap": round(media_diaria_ap, 2),
        "dias_com_entrada_ap": dias_ap,
        "media_semanal": round(total_ap / 4, 2) if total_ap > 0 else 0,
        "semanas_analisadas": 4,
        "entradas_mes_atual": round(entradas_mes, 2),
        "media_diaria_atual": round(entradas_mes / max(dias_com_entrada, 1), 2),
        "dias_restantes": dias_rest,
        "previsao_entrada_restante": round(previsao_entrada_restante, 2),
        "previsao_mes": round(entradas_mes + previsao_entrada_restante, 2),
        "outliers_removidos": round(outliers_removidos, 2),
    }


def simular_prazos_compra(contas):
    """Simula onde cairia o vencimento se liberar pedido hoje com prazos 30/40/50/60/90 dias.
    Mostra a carga de contas que já existe em cada data pra avaliar viabilidade."""
    hoje = date.today()
    prazos = [30, 40, 50, 60, 90]

    # Carga de contas previstas por dia (próximos 90 dias)
    carga_por_dia = {}
    for c in contas:
        if c["status"] == "Previsto" and c["data_pagamento"]:
            d = c["data_pagamento"]
            if hoje < d <= hoje + timedelta(days=95):
                carga_por_dia[d] = carga_por_dia.get(d, 0) + c["valor"]

    # Carga por semana
    carga_por_semana = {}
    for c in contas:
        if c["status"] == "Previsto" and c["data_pagamento"]:
            d = c["data_pagamento"]
            if hoje < d <= hoje + timedelta(days=95):
                sem_inicio = d - timedelta(days=d.weekday())
                carga_por_semana[sem_inicio] = carga_por_semana.get(sem_inicio, 0) + c["valor"]

    nomes = {0: "Seg", 1: "Ter", 2: "Qua", 3: "Qui", 4: "Sex", 5: "Sab", 6: "Dom"}
    nomes_meses = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
                   7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}

    simulacoes = []
    for p in prazos:
        data_venc = hoje + timedelta(days=p)
        while data_venc.weekday() >= 5:
            data_venc += timedelta(days=1)

        carga_dia = carga_por_dia.get(data_venc, 0)
        sem_inicio = data_venc - timedelta(days=data_venc.weekday())
        carga_semana = carga_por_semana.get(sem_inicio, 0)

        if carga_semana > 15000:
            viabilidade = "critico"
            texto = "Semana ja pesada demais"
        elif carga_semana > 8000:
            viabilidade = "atencao"
            texto = "Semana carregada"
        elif carga_dia > 5000:
            viabilidade = "atencao"
            texto = "Dia pesado"
        else:
            viabilidade = "ok"
            texto = "Boa janela"

        simulacoes.append({
            "prazo": p,
            "data_venc": data_venc,
            "dia_semana": nomes.get(data_venc.weekday(), "?"),
            "mes": nomes_meses.get(data_venc.month, "?"),
            "carga_dia": round(carga_dia, 2),
            "carga_semana": round(carga_semana, 2),
            "viabilidade": viabilidade,
            "texto": texto,
        })

    return simulacoes


def recomendar_data_liberacao_pedido(fornecedor_nome: str, prazos_disponiveis: list, contas: list, valor_pedido: float = 0):
    """
    Recomenda a MELHOR data para liberar um pedido a um fornecedor específico.
    
    Analisa todos os prazos disponíveis do fornecedor e recomenda:
    - Qual prazo escolher (30/60/90 dias)
    - Qual data liberar o pedido pra que caia numa semana mais leve
    
    Args:
        fornecedor_nome: Nome do fornecedor
        prazos_disponiveis: Lista com prazos em dias [30, 60, 90]
        contas: Lista de contas a pagar (com datas_pagamento)
        valor_pedido: Valor estimado do pedido (para extra info)
    
    Returns:
        Dict com recomendação detalhada
    """
    hoje = date.today()
    nomes = {0: "Seg", 1: "Ter", 2: "Qua", 3: "Qui", 4: "Sex", 5: "Sab", 6: "Dom"}
    nomes_meses = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
                   7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}
    
    # Carga por semana (próximas 3 meses)
    carga_por_semana = {}
    for c in contas:
        if c.get("status") in ("Previsto", "Pago", "Paga", "Liquidado") and c.get("data_pagamento"):
            d = c["data_pagamento"]
            if hoje <= d <= hoje + timedelta(days=91):
                sem_inicio = d - timedelta(days=d.weekday())
                carga_por_semana[sem_inicio] = carga_por_semana.get(sem_inicio, 0) + c["valor"]
    
    # Analisar cada prazo
    analise_prazos = []
    for prazo in prazos_disponiveis:
        data_venc = hoje + timedelta(days=prazo)
        while data_venc.weekday() >= 5:  # Skip finais de semana
            data_venc += timedelta(days=1)
        
        sem_inicio = data_venc - timedelta(days=data_venc.weekday())
        carga_semana = carga_por_semana.get(sem_inicio, 0)
        
        # Determinar viabilidade
        if carga_semana > 15000:
            viabilidade = "critico"
            emoji = "⛔"
            score = 0
        elif carga_semana > 8000:
            viabilidade = "atencao"
            emoji = "⚠️"
            score = 1
        else:
            viabilidade = "ok"
            emoji = "✅"
            score = 2
        
        analise_prazos.append({
            "prazo": prazo,
            "data_venc": data_venc,
            "dia_semana": nomes.get(data_venc.weekday(), "?"),
            "mes": nomes_meses.get(data_venc.month, "?"),
            "carga_semana": round(carga_semana, 2),
            "viabilidade": viabilidade,
            "emoji": emoji,
            "score": score,
        })
    
    # Encontrar melhor prazo (score mais alto = menos carga)
    mejor = max(analise_prazos, key=lambda x: x["score"]) if analise_prazos else None
    
    if not mejor:
        return {"recomendacao": "⚠️ Sem dados para recomendar", "mejor": None, "analise": analise_prazos}
    
    # Textos contextualizados
    if mejor["viabilidade"] == "critico":
        contexto = f"⛔ CRÍTICO: Semana de {mejor['data_venc'].strftime('%d/%m')} já tem R$ {mejor['carga_semana']:,.2f} previsto. Negocie prazo diferente ou atrase a liberação."
    elif mejor["viabilidade"] == "atencao":
        contexto = f"⚠️ ATENÇÃO: Semana já está carregada (R$ {mejor['carga_semana']:,.2f}). Procure outro prazo se possível."
    else:
        contexto = f"✅ BOA JANELA: Semana de {mejor['data_venc'].strftime('%d/%m')} tem apenas R$ {mejor['carga_semana']:,.2f} — ótimo momento!"
    
    return {
        "recomendacao": contexto,
        "prazo_recomendado": mejor["prazo"],
        "data_vencimento": mejor["data_venc"],
        "viabilidade": mejor["viabilidade"],
        "carga_semana": mejor["carga_semana"],
        "analise_todos_prazos": analise_prazos,
    }


def analisar_pix_vs_boleto(valor: float, prazo_dias: int = 0, margem_produto_pct: float = 40):
    """
    Analisa se vale a pena converter um pagamento de PIX para Boleto.
    
    Considera:
    - Taxa PIX (0%) vs Taxa Boleto (0.40% em média)
    - Ganho financeiro (pode investir em CDB durante o prazo)
    - Fluxo de caixa (precisa de cash imediato?)
    
    Args:
        valor: Valor da compra
        prazo_dias: Quantos dias de prazo o boleto daria (0 = sem prazo)
        margem_produto_pct: Margem % do produto (padrão 40%)
    
    Returns:
        Dict com análise comparativa
    """
    taxa_boleto = 0.004  # 0.4%
    custo_boleto = valor * taxa_boleto
    
    # Taxa de juros CDB (referência: ~9.5% a.a.)
    taxa_cdb_anual = 0.095
    taxa_cdb_diaria = taxa_cdb_anual / 365
    
    # Ganho financeiro por cada dia de prazo (investindo em CDB)
    ganho_juros = valor * taxa_cdb_diaria * prazo_dias if prazo_dias > 0 else 0
    
    # Fluxo imediato PIX vs prazo em Boleto
    resultado_liquido = ganho_juros - custo_boleto  # positivo = vale a pena Boleto
    
    # Contexto de decisão
    if prazo_dias == 0:
        # Sem prazo — qualidade PIX (sem taxa, sem espera)
        recomendacao = "PIX ✅"
        motivo = "Sem prazo de vencimento. PIX é mais ágil — prefira PIX."
        risco = "baixo"
    elif resultado_liquido > 0:
        recomendacao = "Boleto 📊"
        motivo = f"Vale a pena Boleto por {prazo_dias} dias. Ganho financeiro: +R${ganho_juros:,.2f} (CDB) menos taxa (-R${custo_boleto:,.2f}) = R${resultado_liquido:,.2f} positivo."
        risco = "alto" if margem_produto_pct < 15 else "medio" if margem_produto_pct < 25 else "baixo"
    else:
        recomendacao = "PIX ✅"
        motivo = f"Boleto perde -R${abs(resultado_liquido):,.2f}. Prefira PIX — entrada de caixa imediata vale mais."
        risco = "baixo"
    
    return {
        "recomendacao": recomendacao,
        "motivo": motivo,
        "valor": valor,
        "prazo_dias": prazo_dias,
        "taxa_boleto_custo": round(custo_boleto, 2),
        "ganho_cdb_juros": round(ganho_juros, 2),
        "resultado_liquido": round(resultado_liquido, 2),
        "recomendacao_curta": "PIX" if resultado_liquido < 0 or prazo_dias == 0 else "Boleto",
        "risco_margem": risco,
    }


def gerar_plano_fornecedores(contas):
    """Gera plano visual de compras baseado no fornecedores_mes.py."""
    try:
        from fornecedores_mes import FORNECEDORES_MES
    except ImportError:
        return None

    if not FORNECEDORES_MES:
        return None

    hoje = date.today()

    # Carga por semana nos próximos 120 dias
    carga_por_semana = {}
    for c in contas:
        if c["status"] == "Previsto" and c["data_pagamento"]:
            d = c["data_pagamento"]
            if hoje < d <= hoje + timedelta(days=125):
                sem_inicio = d - timedelta(days=d.weekday())
                carga_por_semana[sem_inicio] = carga_por_semana.get(sem_inicio, 0) + c["valor"]

    nomes_dia = {0: "Seg", 1: "Ter", 2: "Qua", 3: "Qui", 4: "Sex", 5: "Sáb", 6: "Dom"}
    nomes_meses = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
                   7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}

    # Buscar histórico de cada fornecedor
    corte = hoje - timedelta(days=180)
    resultado = []

    for forn in FORNECEDORES_MES:
        nome = forn["nome"]
        # Buscar por partes do nome
        partes_busca = nome.upper().replace("(", " ").replace(")", " ").split()
        principal = partes_busca[0] if partes_busca else nome.upper()

        realizadas = [c for c in contas
                      if principal in c["fornecedor"].upper()
                      and c["status"] == "Realizado"
                      and c["data_pagamento"] and c["data_pagamento"] >= corte]
        previstas = [c for c in contas
                     if principal in c["fornecedor"].upper()
                     and c["status"] == "Previsto"
                     and c["data_pagamento"]]

        total_realizado = sum(c["valor"] for c in realizadas)
        media_pedido = total_realizado / max(len(realizadas), 1)
        total_previsto = sum(c["valor"] for c in previstas)

        # Determinar risco
        if forn.get("observacao") and ("risco" in forn["observacao"].lower()
                                        or "ruim" in str(forn.get("observacao", "")).lower()):
            risco = "alto"
        elif not realizadas or forn.get("observacao", "").startswith("Fornecedor novo"):
            risco = "medio"
        elif forn.get("prazo_livre"):
            risco = "baixo"
        else:
            # Verificar se algum prazo cai em mês lotado
            tem_ok = False
            for p in forn["prazos"]:
                d = hoje + timedelta(days=p)
                while d.weekday() >= 5:
                    d += timedelta(days=1)
                sem = d - timedelta(days=d.weekday())
                carga = carga_por_semana.get(sem, 0)
                if carga < 8000:
                    tem_ok = True
                    break
            risco = "baixo" if tem_ok else "medio"

        # Simular prazos
        simulacoes = []
        melhor_prazo = None
        melhor_carga = float("inf")

        for p in forn["prazos"]:
            d = hoje + timedelta(days=p)
            while d.weekday() >= 5:
                d += timedelta(days=1)
            sem = d - timedelta(days=d.weekday())
            carga = carga_por_semana.get(sem, 0)

            if carga > 15000:
                status = "critico"
                texto = "Semana pesada demais"
            elif carga > 8000:
                status = "atencao"
                texto = "Semana carregada"
            else:
                status = "ok"
                texto = "Boa janela"

            sim = {
                "prazo": p,
                "data": d,
                "dia_semana": nomes_dia.get(d.weekday(), "?"),
                "mes": nomes_meses.get(d.month, "?"),
                "carga_semana": round(carga, 2),
                "status": status,
                "texto": texto,
                "recomendado": False,
            }
            simulacoes.append(sim)

            if carga < melhor_carga:
                melhor_carga = carga
                melhor_prazo = sim

        # Marcar recomendados (status ok)
        for s in simulacoes:
            if s["status"] == "ok":
                s["recomendado"] = True

        # AQUi: Análise de pico - se TODOS os prazos vencerem no mesmo dia
        prazos_por_data = {}
        for s in simulacoes:
            d = s["data"]
            if d not in prazos_por_data:
                prazos_por_data[d] = []
            prazos_por_data[d].append(s)
        
        alerta_pico = None
        if len(prazos_por_data) == 1:
            # Todos os prazos caem no mesmo dia - PROBLEMA!
            data_unica = list(prazos_por_data.keys())[0]
            carga_nesse_dia = carga_por_semana.get(data_unica - timedelta(days=data_unica.weekday()), 0)
            alerta_pico = {
                "tipo": "TODOS_MESMO_DIA",
                "data": data_unica,
                "dia_semana": nomes_dia.get(data_unica.weekday(), "?"),
                "carga_dia": carga_nesse_dia,
                "nivel": "CRITICO" if carga_nesse_dia > 15000 else "ATENCAO" if carga_nesse_dia > 8000 else "OK",
                "aviso": f"ATENÇÃO! Todos os prazos ({'/'.join(str(p['prazo']) for p in simulacoes)}d) vencem no mesmo dia ({data_unica.strftime('%d/%m - %A')}). Não é recomendável!"
            }

        resultado.append({
            "nome": nome,
            "prazos_disponiveis": "/".join(str(p) for p in forn["prazos"]),
            "prazo_livre": forn.get("prazo_livre", False),
            "valor_sugerido": forn.get("valor_sugerido", ""),
            "valor_orcado": forn.get("valor_orcado"),
            "observacao": forn.get("observacao", ""),
            "historico_total": round(total_realizado, 2),
            "historico_count": len(realizadas),
            "historico_media": round(media_pedido, 2),
            "ja_previsto": round(total_previsto, 2),
            "risco": risco,
            "simulacoes": simulacoes,
            "melhor_prazo": melhor_prazo,
            "alerta_pico": alerta_pico,  # NOVO: avisar se todos prazos caem no mesmo dia
        })

    # Carga mensal
    carga_mensal = {}
    for sem, v in carga_por_semana.items():
        m = sem.month
        carga_mensal[m] = carga_mensal.get(m, 0) + v

    # Encontrar janela de ouro (semana mais leve nos próximos 60 dias)
    janela_ouro = None
    min_carga = float("inf")
    for i in range(7, 70):
        d = hoje + timedelta(days=i)
        if d.weekday() == 0:  # Segundas
            sem = d
            carga = carga_por_semana.get(sem, 0)
            if carga < min_carga:
                min_carga = carga
                fim_sem = sem + timedelta(days=4)
                janela_ouro = {
                    "inicio": sem,
                    "fim": fim_sem,
                    "carga": round(carga, 2),
                }

    # Calcular alertas de overspending por mês (histórico)
    alertas_overspending = calcular_alerta_overspending_mensal(contas, carga_mensal)

    return {
        "fornecedores": resultado,
        "carga_mensal": {k: round(v, 2) for k, v in sorted(carga_mensal.items())},
        "alertas_overspending": alertas_overspending,
        "janela_ouro": janela_ouro,
        "data_geracao": hoje,
    }


def calcular_alerta_overspending_mensal(contas, carga_mensal_atual):
    """
    Compara carga de compras MENSAL do ano ATUAL vs HISTÓRICO (ano passado).
    Se está acima do histórico, retorna avisos de risco de overspending.
    
    Args:
        contas: Lista de contas (com datas)
        carga_mensal_atual: Dict {mês: valor} do planejamento atual
    
    Returns:
        Dict com avisos de overspending por mês
    """
    hoje = date.today()
    ano_atual = hoje.year
    ano_passado = ano_atual - 1
    
    # Calcular carga mensual do ano passado (mesmos meses)
    carga_ano_passado = {}
    for c in contas:
        if c.get("status") == "Realizado" and c.get("data_pagamento"):
            d = c["data_pagamento"]
            # Se é do ano passado, contar
            if d.year == ano_passado:
                m = d.month
                carga_ano_passado[m] = carga_ano_passado.get(m, 0) + c["valor"]
    
    nomes_meses_full = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }
    
    avisos_mes = {}
    
    # Comparar cada mês
    for mes, carga_atual in carga_mensal_atual.items():
        carga_passada = carga_ano_passado.get(mes, 0)
        
        if carga_passada == 0:
            continue  # Sem histórico
        
        diff = carga_atual - carga_passada
        pct_diff = (diff / carga_passada) * 100 if carga_passada > 0 else 0
        
        # Determinar nível de alerta
        if diff > carga_passada * 0.5:  # > 50% a mais
            nivel = "critico"
            emoji = "🔴"
            titulo = f"⚠️ CRÍTICO — {nomes_meses_full.get(mes, f'Mês {mes}')}"
            alerta = f"Ano passado: R${carga_passada:,.2f} | Hoje: R${carga_atual:,.2f} | Diferença: +{pct_diff:.0f}% 🚨\n" \
                     f"Você ESTÁ EXTRAPOLANDO em compras neste mês! Risco alto de overspending."
            recomendacao = "Reduza novos pedidos. Venda primeiro o estoque existente antes de comprar mais."
        elif diff > 0:  # Algum aumento
            nivel = "atencao"
            emoji = "🟡"
            titulo = f"⚠️ ATENÇÃO — {nomes_meses_full.get(mes, f'Mês {mes}')}"
            alerta = f"Ano passado: R${carga_passada:,.2f} | Hoje: R${carga_atual:,.2f} | Diferença: +{pct_diff:.0f}%\n" \
                     f"Você está acima do histórico. Cuidado para não repetir padrão de overspending."
            recomendacao = "Monitore de perto e reduza novos pedidos se a receita não aumentar."
        else:  # Menor que ano passado
            nivel = "ok"
            emoji = "✅"
            titulo = f"✅ OK — {nomes_meses_full.get(mes, f'Mês {mes}')}"
            alerta = f"Ano passado: R${carga_passada:,.2f} | Hoje: R${carga_atual:,.2f} | Economia: {abs(pct_diff):.0f}%\n" \
                     f"Você está controlando melhor as compras! Parabéns."
            recomendacao = "Continue assim. Mantenha o rigor."
        
        avisos_mes[mes] = {
            "mes": mes,
            "mes_nome": nomes_meses_full.get(mes, f"Mês {mes}"),
            "carga_ano_passado": round(carga_passada, 2),
            "carga_atual": round(carga_atual, 2),
            "diferenca": round(diff, 2),
            "pct_diferenca": round(pct_diff, 1),
            "nivel": nivel,
            "emoji": emoji,
            "titulo": titulo,
            "alerta": alerta,
            "recomendacao": recomendacao,
        }
    
    return avisos_mes


def gerar_bloco_compras(contas, receita_info, saldo_info):
    """Gera o bloco de planejamento de compras (dia 20+)."""
    hoje = date.today()

    # Mês seguinte
    if hoje.month < 12:
        prox_mes = hoje.month + 1
        prox_ano = hoje.year
    else:
        prox_mes = 1
        prox_ano = hoje.year + 1

    nomes_meses = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
                   5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                   9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"}
    nome_prox = nomes_meses.get(prox_mes, "?")

    # Base de faturamento pra calcular teto
    fat_base = receita_info.get("previsao_mes", VENDA_IDEAL)
    if fat_base < VENDA_MINIMA_SEGURA:
        fat_base = VENDA_MINIMA_SEGURA  # Mínimo conservador

    teto_compra = fat_base * (CMV_TETO_PCT / 100)

    # Compras já previstas pro próximo mês
    try:
        inicio_prox = date(prox_ano, prox_mes, 1)
        if prox_mes < 12:
            fim_prox = date(prox_ano, prox_mes + 1, 1) - timedelta(days=1)
        else:
            fim_prox = date(prox_ano, 12, 31)
    except ValueError:
        fim_prox = date(prox_ano, prox_mes, 28)

    compras_previstas = [c for c in contas if c["status"] == "Previsto"
                         and c["data_pagamento"]
                         and inicio_prox <= c["data_pagamento"] <= fim_prox
                         and c["categoria"] == "Custo das Vendas"]
    total_ja_previsto = sum(c["valor"] for c in compras_previstas)
    disponivel = max(0, teto_compra - total_ja_previsto)

    # Ranking de fornecedores do mês (compras realizadas últimos 6 meses)
    corte_hist = hoje - timedelta(days=180)
    hist_fornecedores = {}
    for c in contas:
        if c["status"] == "Realizado" and c["categoria"] == "Custo das Vendas":
            if c["data_pagamento"] and c["data_pagamento"] >= corte_hist:
                f = c["fornecedor"]
                if f not in hist_fornecedores:
                    hist_fornecedores[f] = {"total": 0, "count": 0, "valores": [],
                                            "formas": Counter(), "datas": []}
                hist_fornecedores[f]["total"] += c["valor"]
                hist_fornecedores[f]["count"] += 1
                hist_fornecedores[f]["valores"].append(c["valor"])
                hist_fornecedores[f]["formas"][c["forma_pagamento"]] += 1
                hist_fornecedores[f]["datas"].append(c["data_pagamento"])

    ranking_fornecedores = []
    for f, d in hist_fornecedores.items():
        media = d["total"] / d["count"] if d["count"] > 0 else 0
        forma_principal = d["formas"].most_common(1)[0][0] if d["formas"] else "?"
        mensal = d["total"] / 6  # 6 meses

        # Melhor dia para vencimento (dia do mês mais frequente)
        dias_venc = Counter(dt.day for dt in d["datas"])
        dia_mais_freq = dias_venc.most_common(1)[0][0] if dias_venc else 15

        ranking_fornecedores.append({
            "fornecedor": f, "total_6m": round(d["total"], 2),
            "count": d["count"], "media": round(media, 2),
            "mensal": round(mensal, 2), "forma": forma_principal,
            "dia_vencimento_freq": dia_mais_freq,
        })
    ranking_fornecedores.sort(key=lambda x: x["total_6m"], reverse=True)

    # Melhores dias pra liberar pedido (dias com menos contas a pagar)
    carga_dias_prox = {}
    for c in contas:
        if c["status"] == "Previsto" and c["data_pagamento"]:
            if inicio_prox <= c["data_pagamento"] <= fim_prox:
                dia = c["data_pagamento"].day
                carga_dias_prox[dia] = carga_dias_prox.get(dia, 0) + c["valor"]

    # Semanas do próximo mês — qual é a mais leve?
    semanas_prox = {}
    for c in contas:
        if c["status"] == "Previsto" and c["data_pagamento"]:
            if inicio_prox <= c["data_pagamento"] <= fim_prox:
                sem = (c["data_pagamento"].day - 1) // 7 + 1
                semanas_prox[sem] = semanas_prox.get(sem, 0) + c["valor"]

    # Ponto de equilíbrio
    impostos = fat_base * 0.09
    fixos_total = sum(v for v in CUSTOS_FIXOS_REF.values() if v is not None) + impostos
    lucro_projetado = fat_base - teto_compra - fixos_total

    # CMV real do mês (compras realizadas no mês atual)
    inicio_mes = date(hoje.year, hoje.month, 1)
    compras_mes_real = sum(c["valor"] for c in contas
                          if c["status"] == "Realizado"
                          and c["categoria"] == "Custo das Vendas"
                          and c["data_pagamento"]
                          and inicio_mes <= c["data_pagamento"] <= hoje)
    fat_mes_atual = receita_info.get("entradas_mes_atual", 1)
    cmv_real_pct = (compras_mes_real / fat_mes_atual * 100) if fat_mes_atual > 0 else 0

    return {
        "mes_referencia": f"{nome_prox}/{prox_ano}",
        "fat_base": round(fat_base, 2),
        "teto_compra": round(teto_compra, 2),
        "ja_previsto": round(total_ja_previsto, 2),
        "disponivel": round(disponivel, 2),
        "ranking_fornecedores": ranking_fornecedores[:15],
        "semanas_prox": semanas_prox,
        "lucro_projetado": round(lucro_projetado, 2),
        "cmv_real_pct": round(cmv_real_pct, 1),
        "compras_mes_real": round(compras_mes_real, 2),
        "fixos_total": round(fixos_total, 2),
    }


def gerar_resumao_fechamento(contas, receita_info):
    """Gera resumão do mês para fechamento: fornecedores, CMV, prazos, acertos."""
    hoje = date.today()
    inicio_mes = date(hoje.year, hoje.month, 1)

    # Compras realizadas no mês
    compras_mes = [c for c in contas
                   if c["status"] == "Realizado"
                   and c["categoria"] == "Custo das Vendas"
                   and c["data_pagamento"]
                   and inicio_mes <= c["data_pagamento"] <= hoje]

    total_compras = sum(c["valor"] for c in compras_mes)
    fat_mes = receita_info.get("entradas_mes_atual", 1)
    cmv_pct = (total_compras / fat_mes * 100) if fat_mes > 0 else 0

    # Ranking fornecedores do mês
    forn_mes = {}
    for c in compras_mes:
        f = c["fornecedor"]
        if f not in forn_mes:
            forn_mes[f] = {"total": 0, "count": 0, "forma": Counter()}
        forn_mes[f]["total"] += c["valor"]
        forn_mes[f]["count"] += 1
        forn_mes[f]["forma"][c["forma_pagamento"]] += 1

    ranking = []
    for f, d in forn_mes.items():
        forma = d["forma"].most_common(1)[0][0] if d["forma"] else "?"
        ranking.append({"fornecedor": f, "total": round(d["total"], 2),
                        "count": d["count"], "forma": forma})
    ranking.sort(key=lambda x: x["total"], reverse=True)

    # Saídas totais do mês (todas as categorias)
    saidas_mes = {}
    for c in contas:
        if c["status"] == "Realizado" and c["data_pagamento"]:
            if inicio_mes <= c["data_pagamento"] <= hoje:
                cat = c["categoria"] if c["categoria"] else "Outros"
                saidas_mes[cat] = saidas_mes.get(cat, 0) + c["valor"]

    # Análise: acertos e melhorias
    acertos = []
    melhorias = []

    if cmv_pct <= CMV_TETO_PCT:
        acertos.append(f"CMV controlado em {cmv_pct:.0f}% (teto: {CMV_TETO_PCT:.0f}%)")
    else:
        melhorias.append(f"CMV em {cmv_pct:.0f}% — acima do teto de {CMV_TETO_PCT:.0f}%. "
                        f"Excesso: R${total_compras - fat_mes * CMV_TETO_PCT / 100:,.2f}")

    if fat_mes >= VENDA_IDEAL:
        acertos.append(f"Faturamento acima do ideal (R${fat_mes:,.2f})")
    elif fat_mes >= VENDA_MINIMA_SEGURA:
        acertos.append(f"Faturamento acima do mínimo seguro (R${fat_mes:,.2f})")
    else:
        melhorias.append(f"Faturamento abaixo do mínimo seguro de R${VENDA_MINIMA_SEGURA:,.2f}")

    # Concentração de fornecedores
    if ranking and len(ranking) >= 3:
        top3_total = sum(r["total"] for r in ranking[:3])
        if top3_total / total_compras > 0.6 and total_compras > 0:
            melhorias.append(f"Top 3 fornecedores concentram {top3_total/total_compras*100:.0f}% das compras — diversifique")

    return {
        "total_compras": round(total_compras, 2),
        "cmv_pct": round(cmv_pct, 1),
        "faturamento": round(fat_mes, 2),
        "ranking_fornecedores": ranking[:10],
        "saidas_por_categoria": {k: round(v, 2) for k, v in
                                  sorted(saidas_mes.items(), key=lambda x: x[1], reverse=True)},
        "acertos": acertos,
        "melhorias": melhorias,
    }


def gerar_propostas_lucro(receita_info, saldos_contas, contas):
    """Gera propostas estratégicas pra aumentar lucro baseado no cenário atual."""
    propostas = []
    hoje = date.today()
    entradas_mes = receita_info.get("entradas_mes_atual", 0)
    inicio_mes = date(hoje.year, hoje.month, 1)

    # Compras realizadas no mês
    compras_realizadas = sum(c["valor"] for c in contas
                            if c["status"] == "Realizado"
                            and c["categoria"] == "Custo das Vendas"
                            and c["data_pagamento"]
                            and inicio_mes <= c["data_pagamento"] <= hoje)
    cmv_atual_pct = (compras_realizadas / entradas_mes * 100) if entradas_mes > 0 else 0
    cmv_ideal = entradas_mes * (CMV_TETO_PCT / 100)
    economia_possivel = compras_realizadas - cmv_ideal

    # 1) CMV
    if cmv_atual_pct > CMV_TETO_PCT and economia_possivel > 0:
        propostas.append({
            "icone": "🎯",
            "titulo": "Reduzir CMV para 35%",
            "proposta": f"O CMV está em {cmv_atual_pct:.0f}%. Se reduzir pra {CMV_TETO_PCT:.0f}%, "
                        f"economiza R${economia_possivel:,.2f}/mês. Como: negocie descontos por volume, "
                        f"avalie fornecedores alternativos, compre só o que gira rápido.",
            "impacto": round(economia_possivel, 2),
        })
    elif cmv_atual_pct <= CMV_TETO_PCT:
        propostas.append({
            "icone": "✅",
            "titulo": "CMV controlado!",
            "proposta": f"CMV em {cmv_atual_pct:.0f}% — dentro do teto. Continue assim! "
                        f"Cada 1% a menos = ~R${entradas_mes * 0.01:,.2f} a mais de lucro.",
            "impacto": 0,
        })

    # 2) Giro de estoque
    propostas.append({
        "icone": "📦",
        "titulo": "Girar estoque parado antes de comprar novo",
        "proposta": "Antes de cada pedido novo, levante o que está encalhado. "
                    "Promoção pra girar estoque antigo libera caixa e espaço. "
                    "Regra: só comprar produto novo se o atual girou 70%+.",
        "impacto": 0,
    })

    # 3) Prazos de pagamento — muito Pix?
    compras_mes = [c for c in contas
                   if c["status"] == "Realizado"
                   and c["categoria"] == "Custo das Vendas"
                   and c["data_pagamento"]
                   and inicio_mes <= c["data_pagamento"] <= hoje]
    pix_count = sum(1 for c in compras_mes if c["forma_pagamento"] == "Pix")
    total_count = len(compras_mes)
    if total_count > 0 and pix_count / total_count > 0.4:
        propostas.append({
            "icone": "💳",
            "titulo": "Migrar Pix → Boleto 30/60/90 dias",
            "proposta": f"{pix_count} de {total_count} compras foram no Pix (pagamento imediato). "
                        f"Negociando boleto 30/60/90 dias, o dinheiro fica no caixa mais tempo. "
                        f"Priorize: fornecedores negociáveis (Naturalles, Haskell, Segali).",
            "impacto": 0,
        })

    # 4) Meta Nubank
    nub_falta = saldos_contas.get("nubank_falta_meta", 0)
    if nub_falta > 0:
        semanas_pra_meta = nub_falta / TRANSF_NUBANK_SEMANAL
        propostas.append({
            "icone": "💜",
            "titulo": f"Meta Nubank R$50k — faltam R${nub_falta:,.2f}",
            "proposta": f"Mantendo R${TRANSF_NUBANK_SEMANAL:,.0f}/semana, a meta é atingida em "
                        f"~{semanas_pra_meta:.0f} semanas ({semanas_pra_meta/4:.0f} meses). "
                        f"Nas semanas boas, tente transferir R$1.000 pra acelerar.",
            "impacto": 0,
        })

    # 5) Sazonalidade — preparação inteligente
    sazonal = get_alerta_sazonal()
    prox_forte = [a for a in sazonal if a.get("impacto_compras") in ("alto", "muito alto")]
    if prox_forte:
        p = prox_forte[0]
        propostas.append({
            "icone": "📅",
            "titulo": f"Preparar {p['titulo']}",
            "proposta": f"{p['alerta']} Negocie prazos LONGOS (60/90 dias) "
                        f"pra que o pagamento caia DEPOIS do mês forte, quando o caixa estiver cheio.",
            "impacto": 0,
        })

    # 6) Ticket médio
    propostas.append({
        "icone": "🎁",
        "titulo": "Aumentar ticket médio com combos e kits",
        "proposta": "Cada R$10 a mais por venda, em 200 vendas/mês, são R$2.000 extras de faturamento "
                    "sem custo adicional de marketing. Monte kits atrativos com margem boa.",
        "impacto": 2000,
    })

    return propostas


def sugerir_adiamentos(sem, saldo_info, contas):
    """Sugere melhores datas para adiar boletos negociáveis quando o caixa não cobre."""
    saldo = saldo_info["saldo"]
    total = sem["total_semana"]

    if saldo >= total:
        return []  # Caixa cobre, não precisa adiar

    # Quanto precisa aliviar
    falta = total - saldo
    margem = 1000  # Margem de segurança mínima
    aliviar_necessario = falta + margem

    # Pegar contas negociáveis da semana, ordenadas por valor (maiores primeiro)
    negociaveis = []
    for d in sem["dias"]:
        for c in d["contas"]:
            if c["negociavel"] and not c["fixa"]:
                negociaveis.append(c)
    negociaveis.sort(key=lambda x: x["valor"], reverse=True)

    if not negociaveis:
        return []

    # Calcular carga de contas por dia nas próximas 4 semanas
    fim_janela = sem["fim"] + timedelta(days=30)
    carga_por_dia = {}
    for c in contas:
        if c["status"] == "Previsto" and c["data_pagamento"]:
            if sem["fim"] < c["data_pagamento"] <= fim_janela:
                d = c["data_pagamento"]
                carga_por_dia[d] = carga_por_dia.get(d, 0) + c["valor"]

    # Encontrar os dias mais leves nas próximas 4 semanas (só dias úteis)
    dias_candidatos = []
    for i in range(7, 31):  # De 7 a 30 dias à frente
        d = sem["fim"] + timedelta(days=i)
        if d.weekday() >= 5:  # Pular fim de semana
            continue
        carga = carga_por_dia.get(d, 0)
        dias_candidatos.append({"data": d, "carga": carga})

    dias_candidatos.sort(key=lambda x: x["carga"])

    # Selecionar quais boletos adiar e para quando
    adiamentos = []
    acum_aliviado = 0
    melhor_data_idx = 0

    for c in negociaveis:
        if acum_aliviado >= aliviar_necessario:
            break
        if melhor_data_idx >= len(dias_candidatos):
            break

        # Pegar o dia mais leve disponível
        melhor = dias_candidatos[melhor_data_idx]
        # Não concentrar tudo no mesmo dia — se já mandou boleto pra esse dia, pula
        while melhor_data_idx < len(dias_candidatos) - 1 and melhor["carga"] > 3000:
            melhor_data_idx += 1
            melhor = dias_candidatos[melhor_data_idx]

        adiamentos.append({
            "fornecedor": c["fornecedor"],
            "descricao": c["descricao"],
            "valor": c["valor"],
            "data_original": c["data_pagamento"],
            "data_sugerida": melhor["data"],
            "carga_dia_sugerido": melhor["carga"],
        })

        # Atualizar carga do dia escolhido
        melhor["carga"] += c["valor"]
        acum_aliviado += c["valor"]
        melhor_data_idx += 1

    return adiamentos


def gerar_mensagem_wpp(sem, saldo_info, nubank, meta_ads, adiamentos):
    """Gera mensagem completa formatada para WhatsApp com valores alinhados."""
    partes = []
    NBSP = "\u00A0"  # Espaço não-quebrável (mantém alinhamento no WhatsApp)

    def formatar_linha(desc, valor_str, largura=42):
        """Alinha descrição à esquerda e valor à direita."""
        desc_trunc = desc[:largura]
        espacos = largura - len(desc_trunc)
        if espacos < 1:
            espacos = 1
        return f"{desc_trunc}{NBSP * espacos}{valor_str}"

    def fmt_valor(v):
        """Formata valor como R$ 1.234,56"""
        return f"R$ {v:>9,.2f}"

    partes.append("📋 *CONTAS A PAGAR NA SEMANA*")
    partes.append(f"Semana {sem['inicio'].strftime('%d/%m')} a {sem['fim'].strftime('%d/%m/%Y')}")
    partes.append(f"Saldo Sicoob: {fmt_valor(saldo_info['saldo'])}")
    partes.append("_________")

    saldo_corrente = saldo_info["saldo"]
    for d in sem["dias"]:
        partes.append(f"*{d['nome_dia']} {d['data'].strftime('%d/%m')}*")
        
        # Contas a pagar (Previsto)
        if d["contas"]:
            for c in sorted(d["contas"], key=lambda x: x["valor"], reverse=True):
                conta_tag = f" [{c['conta_financeira']}]" if c["conta_financeira"] not in ("SICOOB", "", "None") else ""
                desc_base = c['descricao'][:38 - len(conta_tag)] + conta_tag
                forn = (c.get("fornecedor") or "")[:25]
                forma = (c.get("forma_pagamento") or "")[:12]
                linha = formatar_linha(desc_base, fmt_valor(c['valor']))
                extra = "  ".join(x for x in [forn, forma] if x)
                if extra:
                    linha += f"  {extra}"
                partes.append(linha)
        
        # Contas pagas (Realizado/Pago)
        if d.get("pagas"):
            partes.append("✅ *JÁ PAGO:*")
            for c in sorted(d["pagas"], key=lambda x: x["valor"], reverse=True):
                conta_tag = f" [{c['conta_financeira']}]" if c["conta_financeira"] not in ("SICOOB", "", "None") else ""
                desc_base = c['descricao'][:38 - len(conta_tag)] + conta_tag
                forn = (c.get("fornecedor") or "")[:25]
                forma = (c.get("forma_pagamento") or "")[:12]
                linha = f"✓ {formatar_linha(desc_base, fmt_valor(c['valor']))}"
                extra = "  ".join(x for x in [forn, forma] if x)
                if extra:
                    linha += f"  {extra}"
                partes.append(linha)
        
        # Apenas mostra zerado se NÃO há contas a pagar NEM pagas
        if not d["contas"] and not d.get("pagas"):
            partes.append("Zerado")
        
        if d.get("pagas"):
            total_pagas_dia = sum(c["valor"] for c in d["pagas"])
            saldo_antes = saldo_corrente
            saldo_corrente -= total_pagas_dia  # Só desconta contas efetivamente pagas
            partes.append(formatar_linha("*Já pago*", fmt_valor(total_pagas_dia)))
            # Exibe fórmula: saldo_antes - total_pagas_dia = saldo_corrente (já calculado acima)
            partes.append(f"*TOTAL NO CAIXA* = R$ {saldo_antes:,.2f} - R$ {total_pagas_dia:,.2f} = R$ {saldo_corrente:,.2f}")
        
        if d["contas"]:
            partes.append(formatar_linha("*TOTAL*", fmt_valor(d['total'])))
        partes.append("_________")

    partes.append(f"\n💰 *TOTAL DA SEMANA:{NBSP * 15}{fmt_valor(sem['total_semana'])}*")

    # ── Decisões ──
    partes.append("\n📌 *DECISÕES DA SEMANA:*")

    # Nubank
    if nubank.get("transferir"):
        partes.append(f"💜 Nubank: Transferir R$ {nubank['valor']:,.2f} na {nubank['dia']}")
    else:
        partes.append("💜 Nubank: ❌ NÃO transferir essa semana (acumula pra próxima)")

    # Meta Ads
    if meta_ads.get("investir"):
        partes.append(f"📣 Tráfego: ✅ Colocar R$ {meta_ads['valor']:,.2f} em campanha")
    else:
        partes.append("📣 Tráfego: ⏸️ Segurar campanhas essa semana")

    # Adiamentos
    if adiamentos:
        partes.append("\n🤝 *NEGOCIAR ADIAMENTO:*")
        for a in adiamentos:
            nome = a["fornecedor"] if a["fornecedor"] != a["descricao"][:50] else a["descricao"][:40]
            partes.append(
                f"  → {nome[:23]:<23s}{NBSP}{fmt_valor(a['valor'])}"
                f"\n{NBSP * 6}vence {a['data_original'].strftime('%d/%m')} → adiar p/ *{a['data_sugerida'].strftime('%d/%m')}*"
            )
        total_adiar = sum(a["valor"] for a in adiamentos)
        partes.append(formatar_linha("  *Total a negociar*", fmt_valor(total_adiar)))

    return "\n".join(partes)


def analisar_distribuicao_rigorosa_fornecedores(contas, prazos_recomendados_por_forn):
    """
    ANÁLISE RIGOROSA: Verifica se múltiplos fornecedores vencendo no mesmo dia
    criam um PICO DE CARGA que inviabiliza o fluxo de caixa.

    Args:
        contas: Lista de todas as contas
        prazos_recomendados_por_forn: Dict {nome_fornecedor: {prazo: X, data: date, ...}}

    Returns:
        Dict com:
        - picos_detectados: List de datas com múltiplos vencimentos
        - carga_por_data: Dict {data: valor_total}
        - recomendacoes_distribuicao: List com alternativas
        - alertas_rigor: List com avisos
    """
    # O código real da tua função começa logo aqui abaixo!
    hoje = date.today()
    
    # Agrupa por data de vencimento
    carga_por_data = {}
    fornecedores_por_data = {}
    
    for forn_nome, rec in prazos_recomendados_por_forn.items():
        d = rec.get("data_vencimento", rec.get("data"))
        if d:
            carga_por_data[d] = carga_por_data.get(d, 0) + rec.get("valor_estimado", 5000)
            if d not in fornecedores_por_data:
                fornecedores_por_data[d] = []
            fornecedores_por_data[d].append(forn_nome)
    
    # Detectar picos (datas com múltiplos fornecedores)
    picos = []
    alertas = []
    
    for data, valor_total in sorted(carga_por_data.items()):
        forts = fornecedores_por_data.get(data, [])
        
        if len(forts) > 1:
            # PICO DETECTADO
            picos.append({
                "data": data,
                "dia": ['Seg','Ter','Qua','Qui','Sex','Sab','Dom'][data.weekday()],
                "num_fornecedores": len(forts),
                "carga_total": valor_total,
                "fornecedores": forts,
            })
            
            # Verificar se é crítico
            if valor_total > 15000:
                alertas.append({
                    "nivel": "CRITICO",
                    "data": data.strftime("%d/%m"),
                    "dia": ['Seg','Ter','Qua','Qui','Sex','Sab','Dom'][data.weekday()],
                    "carga": valor_total,
                    "msg": f"PICO PERIGOSO em {data.strftime('%d/%m')} ({['Seg','Ter','Qua','Qui','Sex','Sab','Dom'][data.weekday()]}): {len(forts)} fornecedores vencendo JUNTOS = R$ {valor_total:,.2f}",
                    "fornecedores": forts,
                })
            elif valor_total > 8000:
                alertas.append({
                    "nivel": "ATENCAO",
                    "data": data.strftime("%d/%m"),
                    "dia": ['Seg','Ter','Qua','Qui','Sex','Sab','Dom'][data.weekday()],
                    "carga": valor_total,
                    "msg": f"Semana de {data.strftime('%d/%m')} com {len(forts)} fornecedores simultaneamente = R$ {valor_total:,.2f}",
                    "fornecedores": forts,
                })
    
    # Gerar recomendações de distribuição
    recomendacoes = []
    if alertas:
        recomendacoes.append({
            "tipo": "distribuicao",
            "msg": "RECOMENDACAO RIGOROSA: Distribua os prazos para espalharnos múltiplos dias da semana.",
            "opcoes": []
        })
        for pico in picos:
            if pico.get("num_fornecedores", 0) > 1:
                recomendacoes[0]["opcoes"].append(
                    f"Opção: Negociar prazos DIFERENTES para {pico['num_fornecedores']} fornecedores em {pico['data'].strftime('%d/%m')}"
                )
    
    return {
        "picos_detectados": picos,
        "carga_por_data": carga_por_data,
        "alertas_rigor": alertas,
        "recomendacoes_distribuicao": recomendacoes,
        "total_picos": len(picos),
    }


def analisar_melhor_dia_liberar_pedido(contas):
    """
    Analisa historicamente qual DIA DA SEMANA é melhor liberar pedidos.
    (Qual dia tem menos carga de contas a pagar?)
    
    Retorna:
        Dict com melhor dia e carga por dia da semana
    """
    nomes_dias = {0: "Seg", 1: "Ter", 2: "Qua", 3: "Qui", 4: "Sex", 5: "Sab", 6: "Dom"}
    carga_por_dia = {i: 0 for i in range(7)}
    
    # Somar carga por dia da semana (próximos 90 dias)
    hoje = date.today()
    for c in contas:
        if c.get("status") in ("Previsto", "Pago", "Paga", "Liquidado") and c.get("data_pagamento"):
            d = c["data_pagamento"]
            if hoje <= d <= hoje + timedelta(days=91):
                dia_semana = d.weekday()
                carga_por_dia[dia_semana] += c["valor"]
    
    # Encontrar melhor dia (menor carga)
    melhor_dia = min(carga_por_dia, key=carga_por_dia.get)
    carga_melhor = carga_por_dia[melhor_dia]
    
    # Análise textual
    dias_ranking = sorted(carga_por_dia.items(), key=lambda x: x[1])
    
    analise = {
        "melhor_dia": melhor_dia,
        "melhor_dia_nome": nomes_dias[melhor_dia],
        "carga_melhor_dia": round(carga_melhor, 2),
        "carga_por_dia": {nomes_dias[k]: round(v, 2) for k, v in carga_por_dia.items()},
        "ranking_dias": [(nomes_dias[d], round(c, 2)) for d, c in dias_ranking],
        "recomendacao": f"Melhor dia para liberar pedido: {nomes_dias[melhor_dia]} (carga: R${carga_melhor:,.2f})",
    }
    
    return analise


def validar_limites_compra(valor_pedido: float, prazo_dias: int, data_vencimento: date) -> dict:
    """
    Valida se um pedido respeita os LIMITES DE COMPRA:
    - Semanal: R$15.000 (máx R$20.000 em emergência)
    - Diário: R$8.000
    - Parcela (se 2+ parcelas): R$2.000 por parcela
    
    Returns:
        Dict com status de validação e avisos
    """
    # Limites
    LIMITE_SEMANAL_IDEAL = 15000
    LIMITE_SEMANAL_EMERGENCIA = 20000
    LIMITE_DIARIO = 8000
    LIMITE_PARCELA = 2000
    
    alertas = []
    warnings = []
    
    # Validação básica
    if valor_pedido <= 0:
        return {"valido": False, "motivo": "Valor inválido", "alertas": [], "warnings": []}
    
    # Limite diário (assumindo que cai em 1 dia)
    if valor_pedido > LIMITE_DIARIO:
        warnings.append({
            "tipo": "diario",
            "nivel": "alerta",
            "mensagem": f"Pedido de R${valor_pedido:,.2f} ultrapassa limite diário de R${LIMITE_DIARIO:,.2f}",
            "sugestao": "Divida em 2 dias de vencimento (próxima semana)",
        })
    
    # Limite semanal (simples check)
    if valor_pedido > LIMITE_SEMANAL_IDEAL:
        if valor_pedido <= LIMITE_SEMANAL_EMERGENCIA:
            warnings.append({
                "tipo": "semanal",
                "nivel": "atencao",
                "mensagem": f"Pedido de R${valor_pedido:,.2f} ultrapassa ideal de R${LIMITE_SEMANAL_IDEAL:,.2f}",
                "sugestao": f"Só em emergência — máximo é R${LIMITE_SEMANAL_EMERGENCIA:,.2f}",
            })
        else:
            warnings.append({
                "tipo": "semanal",
                "nivel": "critico",
                "mensagem": f"Pedido de R${valor_pedido:,.2f} ULTRAPASSA LIMITE de R${LIMITE_SEMANAL_EMERGENCIA:,.2f}!",
                "sugestao": "NÃO LIBERAR. Reduza para R$20.000 máx ou divida em 2 semanas.",
            })
    
    status_ok = len([w for w in warnings if w["nivel"] == "critico"]) == 0
    
    return {
        "valido": status_ok,
        "valor": valor_pedido,
        "limite_semanal_ideal": LIMITE_SEMANAL_IDEAL,
        "limite_semanal_emergencia": LIMITE_SEMANAL_EMERGENCIA,
        "limite_diario": LIMITE_DIARIO,
        "limite_parcela": LIMITE_PARCELA,
        "avisos": alertas,
        "warnings": warnings,
    }


def propor_divisao_parcelas(valor_pedido: float, prazo_dias: int = 60) -> dict:
    """
    Se parcela > R$2.000, propõe dividir em N+1 parcelas.
    Retorna sugestão de parcelamento.
    
    Args:
        valor_pedido: Valor total do pedido
        prazo_dias: Quantos dias de prazo tem
    
    Returns:
        Dict com proposta de parcelamento
    """
    LIMITE_PARCELA = 2000
    
    # Se cabe em 1 parcela, OK
    if valor_pedido <= LIMITE_PARCELA:
        return {
            "prazos_sugeridos": [prazo_dias],
            "parcelas": [{"numero": 1, "valor": valor_pedido, "dias": prazo_dias}],
            "precisa_dividir": False,
            "motivo": f"Valor R${valor_pedido:,.2f} <= R${LIMITE_PARCELA:,.2f}",
        }
    
    # Calcular quantas parcelas precisa
    num_parcelas = (valor_pedido / LIMITE_PARCELA) + 1
    num_parcelas = int(num_parcelas) + (1 if valor_pedido % LIMITE_PARCELA > 0 else 0)
    num_parcelas = max(2, num_parcelas)  # Mínimo 2
    
    # Dividir valor
    valor_por_parcela = valor_pedido / num_parcelas
    
    # Distribuir dias (espalhado ao longo do prazo)
    dias_entre_parcelas = max(7, prazo_dias // (num_parcelas - 1)) if num_parcelas > 1 else prazo_dias
    
    parcelas = []
    for i in range(num_parcelas):
        dias_venc = prazo_dias + (i * dias_entre_parcelas)
        parcelas.append({
            "numero": i + 1,
            "valor": round(valor_por_parcela, 2),
            "dias": dias_venc,
            "data_vencimento": date.today() + timedelta(days=dias_venc),
        })
    
    # Prazos sugeridos
    prazos = [p["dias"] for p in parcelas]
    
    return {
        "valor_original": valor_pedido,
        "num_parcelas": num_parcelas,
        "valor_por_parcela": round(valor_por_parcela, 2),
        "limit_parcela": LIMITE_PARCELA,
        "prazos_sugeridos": prazos,
        "parcelas": parcelas,
        "precisa_dividir": True,
        "motivo": f"Valor R${valor_pedido:,.2f} ultrapassa R${LIMITE_PARCELA:,.2f}/parcela",
        "sugestao": f"Divida em {num_parcelas} parcelas de ~R${valor_por_parcela:,.2f}",
    }


def avisar_overspending_historico(receita_info):
    """
    AVISA se ANO PASSADO você extrapolou em compras.
    Detecta qual mês foi mais arriscado.
    
    Returns:
        Dict com warnings de overspending histórico
    """
    avisos = []
    
    # Exemplo de dados históricos (seria lido do Excel)
    # Por enquanto, retornamos estrutura de aviso
    
    avisos.append({
        "titulo": "⚠️ Análise Histórica de Overspending",
        "icone": "🔍",
        "alerta": "Ao carregar o arquivo financeiro, analisaremos se algum mês do ano anterior ultrapassou os limites de CMV (35%) e compras (R$20k/semana).",
        "tipo": "info",
    })
    
    return avisos


def gerar_recomendacoes_compras(contas):
    """
    Gera recomendações inteligentes de QUANDO E COMO liberar pedidos.
    
    Para cada fornecedor em fornecedores_mes.py:
    - Recomenda melhor data pra liberar (qual prazo, qual exato dia)
    - Analisa PIX vs Boleto (vale a pena prazo extra?)
    - Gera card visual com recomendações
    
    Returns:
        List[Dict] com recomendações por fornecedor
    """
    try:
        from fornecedores_mes import FORNECEDORES_MES
    except ImportError:
        return []
    
    if not FORNECEDORES_MES:
        return []
    
    recomendacoes = []
    
    for forn in FORNECEDORES_MES:
        nome = forn.get("nome", "?")
        prazos = forn.get("prazos", [])
        valor_orcado = forn.get("valor_orcado")
        
        # Pular se sem prazos ou dados incompletos
        if not prazos or not nome:
            continue
        
        # Recomendar melhor data
        rec_data = recomendar_data_liberacao_pedido(
            fornecedor_nome=nome,
            prazos_disponiveis=prazos,
            contas=contas,
            valor_pedido=valor_orcado or 5000  # Default 5k se sem info
        )
        
        # Analisar PIX vs Boleto para o melhor prazo
        melhor_prazo = rec_data.get("prazo_recomendado", prazos[0])
        rec_pix_boleto = analisar_pix_vs_boleto(
            valor=valor_orcado or 5000,
            prazo_dias=melhor_prazo,
            margem_produto_pct=40
        )
        
        # Determinar icone de risco
        if rec_data.get("viabilidade") == "critico":
            icone_risco = "🔴"
            urgencia = "CRÍTICO"
        elif rec_data.get("viabilidade") == "atencao":
            icone_risco = "🟡"
            urgencia = "ATENÇÃO"
        else:
            icone_risco = "🟢"
            urgencia = "OK"
        
        recomendacoes.append({
            "fornecedor": nome,
            "icone": icone_risco,
            "urgencia": urgencia,
            "melhor_prazo": melhor_prazo,
            "data_vencimento": rec_data.get("data_vencimento"),
            "carga_semana": rec_data.get("carga_semana"),
            "contexto": rec_data.get("recomendacao"),
            "analise_prazos": rec_data.get("analise_todos_prazos", []),
            "pix_vs_boleto": rec_pix_boleto,
            "recomendacao_pagamento": rec_pix_boleto.get("recomendacao"),
            "motivo_pagamento": rec_pix_boleto.get("motivo"),
        })
    
    return recomendacoes


def carregar_financeiro(fechamento=False):
    """Carrega e processa a planilha financeira completa."""
    caminho = encontrar_planilha()
    if not caminho:
        return None, "Planilha financeira não encontrada no Desktop."

    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
    except Exception as e:
        return None, f"Erro ao abrir planilha: {e}"

    contas = ler_contas_a_pagar(wb)
    saldo = ler_saldo_caixa(wb)
    saldos_contas = ler_saldos_contas(wb)
    semana = gerar_contas_semana(contas)
    projecao = gerar_projecao_caixa(saldo, contas)

    # Previsão de receita (referência: mesmo mês ano anterior)
    receita_info = prever_receita_semanal(wb)

    # Ajustar projeção com receita prevista (referência ano anterior)
    projecao["receita_prevista"] = receita_info.get("previsao_mes", 0)
    previsao_entrada_restante = receita_info.get("previsao_entrada_restante", 0)
    projecao["saldo_com_receita"] = round(projecao["saldo_projetado"] + previsao_entrada_restante, 2)
    projecao["previsao_entrada_restante"] = round(previsao_entrada_restante, 2)

    # Saldo progressivo (caixa - contas do dia, dia a dia)
    saldo_apos_contas = saldo["saldo"] - semana["total_semana"]

    # Decisões inteligentes — baseadas no saldo REAL (sem previsão!)
    entrada_semanal_prevista = receita_info.get("media_semanal", 0)

    insights = gerar_insights_semana(semana, saldo, contas)

    # Calcular teto diário de pagamento baseado na receita
    teto_diario = calcular_teto_diario_pagamento(receita_info, semana)

    # Insight de previsão — APENAS INFORMATIVO (referência ano anterior)
    if receita_info.get("ref_ano_passado", 0) > 0:
        insights.insert(0, {
            "tipo": "info",
            "icone": "📊",
            "texto": f"Referência {receita_info['ref_mes']}: o mês faturou R${receita_info['ref_ano_passado']:,.2f} "
                     f"(~R${receita_info['media_diaria_ap']:,.2f}/dia em {receita_info['dias_com_entrada_ap']} dias). "
                     f"Use como comparativo, NÃO como garantia."
        })

    nubank = calcular_transferencia_nubank(saldo_apos_contas)
    meta_ads = calcular_meta_ads(saldo_apos_contas - (nubank["valor"] if nubank["transferir"] else 0))

    # Avaliação do Nubank (reserva)
    nubank_avaliacao = avaliar_uso_nubank(saldos_contas, saldo_apos_contas)

    # Sugestões de adiamento — SEMPRE baseadas no saldo REAL (sem previsão)
    # A previsão é só informativa, não pode mudar decisões de adiamento
    adiamentos = sugerir_adiamentos(semana, saldo, contas)

    mensagem_wpp = gerar_mensagem_wpp(semana, saldo, nubank, meta_ads, adiamentos)

    # Alertas sazonais
    alertas_sazonais = get_alerta_sazonal()

    # Bloco de compras (dia 20+)
    bloco_compras = None
    hoje = date.today()
    if hoje.day >= 20:
        bloco_compras = gerar_bloco_compras(contas, receita_info, saldo)

    # Resumão fechamento
    resumao = None
    if fechamento:
        resumao = gerar_resumao_fechamento(contas, receita_info)

    # Propostas estratégicas pra aumentar lucro (sempre aparece)
    propostas = gerar_propostas_lucro(receita_info, saldos_contas, contas)

    # Simulação de prazos de compra (30/40/50/60/90 dias)
    simulacao_prazos = simular_prazos_compra(contas)

    # Plano de fornecedores (se fornecedores_mes.py existir)
    plano_fornecedores = gerar_plano_fornecedores(contas)

    # Recomendações de compras: melhor data + PIX vs Boleto
    recomendacoes_compras = gerar_recomendacoes_compras(contas)

    # Análise de melhor dia pra liberar pedidos (historicamente)
    analise_melhor_dia = analisar_melhor_dia_liberar_pedido(contas)

    # Avisos sobre overspending histórico
    avisos_overspending = avisar_overspending_historico(receita_info)

    return {
        "caminho": caminho,
        "contas": contas,
        "saldo": saldo,
        "saldos_contas": saldos_contas,
        "semana": semana,
        "teto_diario": teto_diario,  # Novo: recomendação de teto diário
        "projecao": projecao,
        "receita_info": receita_info,
        "insights": insights,
        "nubank": nubank,
        "nubank_avaliacao": nubank_avaliacao,
        "meta_ads": meta_ads,
        "adiamentos": adiamentos,
        "mensagem_wpp": mensagem_wpp,
        "alertas_sazonais": alertas_sazonais,
        "bloco_compras": bloco_compras,
        "resumao": resumao,
        "propostas": propostas,
        "simulacao_prazos": simulacao_prazos,
        "plano_fornecedores": plano_fornecedores,
        "recomendacoes_compras": recomendacoes_compras,  # Novo: quando liberar pedido + PIX vs Boleto
        "analise_melhor_dia": analise_melhor_dia,  # Novo: qual dia da semana é melhor liberar
        "avisos_overspending": avisos_overspending,  # Novo: warnings sobre overspending histórico
    }, None
